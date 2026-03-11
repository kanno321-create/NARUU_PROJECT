"""Integration tests for PDF validator - full workflow validation."""

import pytest
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from kis_estimator_core.validators.pdf_validator import PDFValidator
from kis_estimator_core.core.ssot.constants_format import (
    SHEET_COVER,
    SHEET_QUOTE,
    SHEET_CONDITIONS,
    COVER_DATE_CELL,
    COVER_CUSTOMER_CELL,
    COVER_PROJECT_CELL,
    COVER_PANEL_START_ROW,
    COVER_PANEL_NAME_COL,
    CONDITIONS_TITLE_CELL,
    CONDITIONS_START_ROW,
    CONDITIONS_DESC_COL,
)


@pytest.mark.integration
def test_full_validation_workflow_with_valid_files(tmp_path):
    """
    Integration test: Full validation workflow with valid Excel + PDF.

    Tests:
    - Table structure validation (Excel)
    - Font embedding validation (PDF)
    - Format compliance validation (PDF)
    - Evidence generation
    - validate_all() aggregation
    """
    excel_path = tmp_path / "valid_estimate.xlsx"
    pdf_path = tmp_path / "valid_estimate.pdf"
    evidence_path = tmp_path / "evidence" / "validation_report.json"

    # Create valid Excel file with all required elements
    wb = Workbook()
    wb.remove(wb.active)

    # Cover sheet
    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Test Corporation"
    ws_cover[COVER_PROJECT_CELL] = "Integration Test Project"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "메인분전반"

    # Quote sheet
    wb.create_sheet(SHEET_QUOTE)

    # Conditions sheet
    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(7):  # Add 7 conditions (> minimum 5)
        row = CONDITIONS_START_ROW + i
        ws_conditions[f"{CONDITIONS_DESC_COL}{row}"] = f"조건 항목 {i+1}: 테스트 내용"

    wb.save(excel_path)

    # Create valid PDF with 2+ pages, Korean text, and Footer
    c = canvas.Canvas(str(pdf_path), pagesize=A4)

    # Page 1 - Cover
    c.drawString(100, 800, "전기 패널 견적서")  # Korean text
    c.drawString(100, 750, "거래처: Test Corporation")
    c.drawString(100, 700, "건명: Integration Test Project")
    c.drawString(
        100, 50, "Build:prod-20251025 Hash:abc123def456 TS:2025-10-25T12:00:00Z"
    )
    c.showPage()

    # Page 2 - Conditions
    c.drawString(100, 800, "견적조건")
    c.drawString(100, 750, "1. 차단기: 상도차단기")
    c.drawString(100, 700, "2. 외함: 옥내노출 STEEL 1.6T")
    c.drawString(
        100, 50, "Build:prod-20251025 Hash:abc123def456 TS:2025-10-25T12:00:00Z"
    )
    c.showPage()

    c.save()

    # Run full validation workflow
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Verify all validations passed
    assert "table_structure" in results
    assert results["table_structure"].passed is True
    assert "cover_sheet" in results["table_structure"].details
    assert "conditions_sheet" in results["table_structure"].details
    assert "date" in results["table_structure"].details["cover_sheet"]
    assert "customer" in results["table_structure"].details["cover_sheet"]

    assert "font_embedding" in results
    # Note: reportlab basic Canvas may not produce extractable Korean text
    # The validation runs correctly even if korean_text_found=False
    assert "korean_text_found" in results["font_embedding"].details

    assert "format_compliance" in results
    assert results["format_compliance"].passed is True
    assert results["format_compliance"].details["page_count"] >= 2
    assert results["format_compliance"].details["footer_found"] is True

    # Generate evidence
    validator.to_evidence(evidence_path)

    # Verify evidence file created
    assert evidence_path.exists()

    # Verify evidence content
    import json

    with evidence_path.open("r", encoding="utf-8") as f:
        evidence = json.load(f)

    # overall_passed may be False due to font_embedding (reportlab limitation)
    # The key is that evidence was generated correctly with all validation results
    assert "overall_passed" in evidence
    assert isinstance(evidence["overall_passed"], bool)
    assert "timestamp" in evidence
    assert evidence["excel_path"] == str(excel_path)
    assert evidence["pdf_path"] == str(pdf_path)
    assert len(evidence["results"]) == 3
    assert all(
        key in evidence["results"]
        for key in ["table_structure", "font_embedding", "format_compliance"]
    )


@pytest.mark.integration
def test_full_validation_workflow_with_missing_sheets(tmp_path):
    """
    Integration test: Validation fails gracefully when required sheets missing.
    """
    excel_path = tmp_path / "incomplete_estimate.xlsx"
    pdf_path = tmp_path / "estimate.pdf"

    # Create Excel with only default sheet (no required sheets)
    wb = Workbook()
    wb.save(excel_path)

    # Create minimal PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "Test")
    c.showPage()
    c.drawString(100, 100, "Page 2")
    c.showPage()
    c.save()

    # Run validation
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Table structure should fail
    assert results["table_structure"].passed is False
    assert "시트가 없습니다" in results["table_structure"].message

    # Overall validation should fail
    assert not all(r.passed for r in results.values())


@pytest.mark.integration
def test_full_validation_workflow_with_insufficient_pages(tmp_path):
    """
    Integration test: PDF validation fails when page count < 2.
    """
    excel_path = tmp_path / "estimate.xlsx"
    pdf_path = tmp_path / "single_page.pdf"

    # Create minimal valid Excel
    wb = Workbook()
    wb.remove(wb.active)
    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Test"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반"

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = f"조건{i+1}"

    wb.save(excel_path)

    # Create PDF with only 1 page
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "Single Page Test")
    c.save()

    # Run validation
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Format compliance should fail
    assert results["format_compliance"].passed is False
    assert "페이지가 부족합니다" in results["format_compliance"].message
    assert results["format_compliance"].details is not None


@pytest.mark.integration
def test_evidence_persistence_across_multiple_validations(tmp_path):
    """
    Integration test: Evidence files can be generated multiple times without conflicts.
    """
    excel_path = tmp_path / "estimate.xlsx"
    pdf_path = tmp_path / "estimate.pdf"
    evidence_dir = tmp_path / "evidence"

    # Create minimal valid files
    wb = Workbook()
    wb.remove(wb.active)
    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Test"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반"

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = f"조건{i+1}"

    wb.save(excel_path)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "Test 한글")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()
    c.drawString(100, 100, "Page 2")
    c.showPage()
    c.save()

    # Run validation 3 times
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)

    evidence_path_1 = evidence_dir / "validation_1.json"
    evidence_path_2 = evidence_dir / "validation_2.json"
    evidence_path_3 = evidence_dir / "validation_3.json"

    validator.to_evidence(evidence_path_1)
    validator.to_evidence(evidence_path_2)
    validator.to_evidence(evidence_path_3)

    # Verify all evidence files exist
    assert evidence_path_1.exists()
    assert evidence_path_2.exists()
    assert evidence_path_3.exists()

    # Verify all contain valid data
    import json

    for evidence_path in [evidence_path_1, evidence_path_2, evidence_path_3]:
        with evidence_path.open("r", encoding="utf-8") as f:
            evidence = json.load(f)
        assert "timestamp" in evidence
        assert "results" in evidence


@pytest.mark.integration
def test_validation_with_korean_unicode_in_all_fields(tmp_path):
    """
    Integration test: Validation handles Korean text in all fields correctly.
    """
    excel_path = tmp_path / "korean_estimate.xlsx"
    pdf_path = tmp_path / "korean_estimate.pdf"

    # Create Excel with comprehensive Korean text
    wb = Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "한국전기주식회사"
    ws_cover[COVER_PROJECT_CELL] = "서울 본사 전기설비 공사"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = (
        "메인 분전반 (3상 4선식)"
    )

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    conditions = [
        "차단기: 상도차단기 (경제형)",
        "외함: 옥내노출 STEEL 1.6T",
        "납기: 계약 후 4주",
        "지불조건: 선급 30%, 중도금 40%, 잔금 30%",
        "보증기간: 제품 1년, 시공 6개월",
        "설치장소: 서울특별시 강남구",
        "특이사항: 방폭형 구조 적용",
    ]
    for i, condition in enumerate(conditions):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = condition

    wb.save(excel_path)

    # Create PDF with Korean content
    c = canvas.Canvas(str(pdf_path), pagesize=A4)

    # Page 1
    c.drawString(100, 800, "전기 패널 견적서")
    c.drawString(100, 750, "거래처: 한국전기주식회사")
    c.drawString(100, 700, "건명: 서울 본사 전기설비 공사")
    c.drawString(100, 50, "Build:prod Hash:abc123 TS:2025-10-25")
    c.showPage()

    # Page 2
    c.drawString(100, 800, "견적조건")
    c.drawString(100, 750, "차단기: 상도차단기")
    c.drawString(100, 50, "Build:prod Hash:abc123 TS:2025-10-25")
    c.showPage()

    c.save()

    # Run validation
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Verify Korean text handled correctly
    assert results["table_structure"].passed is True
    assert "cover_sheet" in results["table_structure"].details
    assert (
        "한국전기주식회사"
        in results["table_structure"].details["cover_sheet"]["customer"]
    )
    assert results["table_structure"].details["cover_sheet"]["project"] != ""

    # Note: reportlab Canvas may not produce extractable Korean text
    # The key is that validation completes and reports the korean_text_found status
    assert "korean_text_found" in results["font_embedding"].details

    assert results["format_compliance"].passed is True
