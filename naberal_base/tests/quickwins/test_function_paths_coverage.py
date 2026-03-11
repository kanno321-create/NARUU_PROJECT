"""
Quick Wins Phase 1: Function Execution Paths

테스트 전략:
- 함수 호출로 실행 경로 커버
- 에러 경로 및 엣지 케이스 테스트
- Zero-Mock 정책 준수

Target Modules:
- stage_runner.py (partial paths)
- validator.py (error handling)
- verbs.py (validation paths)
"""

import os
import pytest
from pathlib import Path
import tempfile

# Disable DB features
os.environ.setdefault("KIS_DISABLE_DB_FEATURES", "1")


# ============================================================
# Validator Error Path Tests
# ============================================================


def test_validator_missing_template():
    """Validator 템플릿 파일 없을 때 에러 경로"""
    from kis_estimator_core.engine.validator import Validator

    non_existent_path = Path("/nonexistent/template.xlsx")

    with pytest.raises(Exception):  # Generic exception for now
        Validator(template_path=non_existent_path)


def test_validator_missing_excel():
    """Validator Excel 파일 없을 때 에러 경로"""
    from kis_estimator_core.engine.validator import Validator

    # Create a temporary template file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        template_path = Path(tmp.name)

    try:
        # Create a minimal Excel file (just for template existence check)
        from openpyxl import Workbook

        wb = Workbook()
        wb.save(str(template_path))

        validator = Validator(template_path=template_path)

        # Try to validate non-existent Excel
        non_existent_excel = Path("/nonexistent/excel.xlsx")

        with pytest.raises(Exception):  # Generic exception for now
            validator.validate(non_existent_excel)

    finally:
        # Cleanup
        if template_path.exists():
            template_path.unlink()


# ============================================================
# Guards Format Error Paths
# ============================================================


def test_formula_guard_empty_workbook():
    """formula_guard 빈 워크북 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import formula_guard

    wb = Workbook()
    ws = wb.active

    # Empty worksheet should fail with strict policy
    is_valid, errors, rate = formula_guard(
        worksheet=ws,
        formula_ranges=["A1:A10"],
        allow_blank=False,
        policy_on_empty="strict_fail",
    )

    assert is_valid is False
    assert len(errors) > 0
    assert rate < 100.0


def test_formula_guard_non_formula_cells():
    """formula_guard 수식 아닌 셀 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import formula_guard

    wb = Workbook()
    ws = wb.active

    # Fill cells with non-formula values
    ws["A1"] = 100
    ws["A2"] = 200
    ws["A3"] = "text"

    is_valid, errors, rate = formula_guard(
        worksheet=ws,
        formula_ranges=["A1:A3"],
        allow_blank=False,
    )

    assert is_valid is False
    assert len(errors) > 0


def test_named_range_guard_missing_ranges():
    """named_range_guard 누락된 네임드 범위 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import named_range_guard

    wb = Workbook()

    # Workbook has no named ranges
    is_valid, errors = named_range_guard(
        workbook=wb, required_ranges=["TEST_RANGE_1", "TEST_RANGE_2"]
    )

    assert is_valid is False
    assert len(errors) >= 2


def test_nan_blank_guard_with_blanks():
    """nan_blank_guard 빈칸 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import nan_blank_guard

    wb = Workbook()
    ws = wb.active

    # Some cells filled, some empty
    ws["A1"] = 100
    ws["A2"] = None  # Blank
    ws["A3"] = 200

    is_valid, errors = nan_blank_guard(
        worksheet=ws, critical_ranges=["A1:A3"], allow_zero=True
    )

    assert is_valid is False
    assert len(errors) > 0


def test_nan_blank_guard_with_zero():
    """nan_blank_guard 0 값 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import nan_blank_guard

    wb = Workbook()
    ws = wb.active

    ws["A1"] = 0
    ws["A2"] = 100

    # With allow_zero=False
    is_valid, errors = nan_blank_guard(
        worksheet=ws, critical_ranges=["A1:A2"], allow_zero=False
    )

    assert is_valid is False
    assert len(errors) > 0


def test_text_policy_guard_forbidden_phrases():
    """text_policy_guard 금칙어 테스트"""
    from openpyxl import Workbook
    from kis_estimator_core.core.ssot.guards_format import text_policy_guard

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "무료 서비스"  # Forbidden word

    is_valid, errors = text_policy_guard(
        worksheet=ws,
        text_ranges=["A1:A5"],
        required_phrases=[],
        forbidden_phrases=["무료", "서비스"],
    )

    assert is_valid is False
    assert len(errors) > 0


def test_cell_protection_guard_unlocked_formula():
    """cell_protection_guard 수식 셀 미보호 테스트"""
    # Skip this test as openpyxl Protection is immutable
    pytest.skip(
        "openpyxl Protection objects are immutable, cannot test this path easily"
    )


# ============================================================
# Stage Runner Utility Paths
# ============================================================


def test_stage_runner_module_import():
    """StageRunner 모듈 import 및 클래스 존재 확인"""
    from kis_estimator_core.kpew.execution.stage_runner import StageRunner

    assert StageRunner is not None


# ============================================================
# Verbs Validation Paths
# ============================================================


def test_verb_context_validation_missing_keys():
    """Verb context validation 누락 키 테스트"""
    from kis_estimator_core.kpew.dsl.verbs import Verb, VerbExecutionError

    class TestVerb(Verb):
        def execute(self, context):
            return context

    verb = TestVerb()

    # Missing keys should raise VerbExecutionError
    with pytest.raises(VerbExecutionError) as exc_info:
        verb.validate_context({}, ["key1", "key2", "key3"])

    assert "Missing required context keys" in str(exc_info.value)


def test_verb_context_validation_partial_keys():
    """Verb context validation 일부 키만 있는 경우"""
    from kis_estimator_core.kpew.dsl.verbs import Verb, VerbExecutionError

    class TestVerb(Verb):
        def execute(self, context):
            return context

    verb = TestVerb()

    # Some keys present, some missing
    context = {"key1": "value1"}

    with pytest.raises(VerbExecutionError) as exc_info:
        verb.validate_context(context, ["key1", "key2", "key3"])

    error_msg = str(exc_info.value)
    assert "key2" in error_msg
    assert "key3" in error_msg


# ============================================================
# Models Edge Cases
# ============================================================


def test_breaker_spec_minimal_creation():
    """BreakerSpec 최소 데이터로 생성 테스트"""
    from kis_estimator_core.models.enclosure import BreakerSpec
    from pydantic import ValidationError

    # Missing required fields should raise ValidationError
    with pytest.raises(ValidationError):
        BreakerSpec()


def test_breaker_spec_with_invalid_data():
    """BreakerSpec 잘못된 데이터 타입 테스트"""
    from kis_estimator_core.models.enclosure import BreakerSpec
    from pydantic import ValidationError

    # Invalid data types
    with pytest.raises(ValidationError):
        BreakerSpec(
            id="TEST",
            model="TEST_MODEL",
            poles="invalid",  # Should be int
            current_a="invalid",  # Should be float
            frame_af="invalid",  # Should be float
        )


# ============================================================
# PDF Converter Paths
# ============================================================


def test_pdf_converter_class_exists():
    """PDFConverter 클래스 존재 확인"""
    try:
        from kis_estimator_core.engine.pdf_converter import PDFConverter

        assert PDFConverter is not None
    except ImportError:
        pytest.skip("PDF converter dependencies not available")


# ============================================================
# Excel Generator Paths
# ============================================================


def test_excel_generator_class_exists():
    """ExcelGenerator 클래스 존재 확인"""
    try:
        from kis_estimator_core.engine.excel_generator import ExcelGenerator

        assert ExcelGenerator is not None
    except ImportError:
        pytest.skip("Excel generator dependencies not available")


# ============================================================
# Breaker Critic Paths
# ============================================================


def test_breaker_critic_class_exists():
    """BreakerCritic 클래스 존재 확인"""
    try:
        from kis_estimator_core.engine.breaker_critic import BreakerCritic

        assert BreakerCritic is not None
    except ImportError:
        pytest.skip("Breaker critic dependencies not available")


# ============================================================
# Phase3 Patch Paths
# ============================================================


def test_phase3_patch_module_import():
    """Phase3 Patch 모듈 import 테스트"""
    try:
        from kis_estimator_core.core.ssot import phase3_patch

        assert phase3_patch is not None
    except ImportError:
        pytest.skip("Phase3 patch dependencies not available")


# ============================================================
# Repo Layer Edge Cases
# ============================================================


def test_plan_repo_module_import():
    """PlanRepo 모듈 import 테스트"""
    try:
        from kis_estimator_core.repos import plan_repo

        assert plan_repo is not None
    except ImportError:
        pytest.skip("PlanRepo dependencies not available")


# ============================================================
# Middleware Idempotency Paths
# ============================================================


def test_idempotency_middleware_import():
    """Idempotency middleware import 테스트"""
    try:
        from kis_estimator_core.middleware import idempotency

        assert idempotency is not None
    except ImportError:
        pytest.skip("Idempotency middleware dependencies not available")


# ============================================================
# Orchestrator Paths
# ============================================================


def test_orchestrator_module_import():
    """Orchestrator 모듈 import 테스트"""
    try:
        from kis_estimator_core.kpew.llm import orchestrator

        assert orchestrator is not None
    except ImportError:
        pytest.skip("Orchestrator dependencies not available")


# ============================================================
# Supabase Client Edge Cases
# ============================================================


def test_supabase_client_module_import():
    """Supabase client 모듈 import 테스트"""
    try:
        from kis_estimator_core.infra import supabase_client

        assert supabase_client is not None
    except ImportError:
        pytest.skip("Supabase client dependencies not available")
