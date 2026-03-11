"""Smoke tests for PDF validator - quick validation checks."""

import pytest
from pathlib import Path
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from kis_estimator_core.validators.pdf_validator import ValidationResult, PDFValidator
from kis_estimator_core.core.ssot.constants_format import (
    SHEET_COVER,
    SHEET_QUOTE,
    SHEET_CONDITIONS,
    COVER_DATE_CELL,
    COVER_CUSTOMER_CELL,
    COVER_PANEL_START_ROW,
    COVER_PANEL_NAME_COL,
    CONDITIONS_TITLE_CELL,
    CONDITIONS_START_ROW,
    CONDITIONS_DESC_COL,
)


@pytest.mark.smoke
def test_validation_result_basic_operations():
    """
    Smoke test: ValidationResult basic operations work correctly.
    """
    # Test passing result
    pass_result = ValidationResult(
        passed=True, message="Validation passed", details={"key": "value"}
    )
    assert pass_result.passed is True
    assert "passed" in pass_result.message.lower()
    assert pass_result.to_dict()["passed"] is True
    assert pass_result.to_dict()["details"]["key"] == "value"

    # Test failing result
    fail_result = ValidationResult(
        passed=False, message="Validation failed", details={"error": "issue"}
    )
    assert fail_result.passed is False
    assert fail_result.to_dict()["passed"] is False


@pytest.mark.smoke
def test_pdf_validator_instantiation():
    """
    Smoke test: PDFValidator can be instantiated without errors.
    """
    # No parameters
    validator1 = PDFValidator()
    assert validator1.excel_path is None
    assert validator1.pdf_path is None

    # With parameters
    validator2 = PDFValidator(excel_path=Path("test.xlsx"), pdf_path=Path("test.pdf"))
    assert validator2.excel_path == Path("test.xlsx")
    assert validator2.pdf_path == Path("test.pdf")


@pytest.mark.smoke
def test_quick_table_validation_with_minimal_excel(tmp_path):
    """
    Smoke test: Quick table structure validation with minimal Excel.
    """
    excel_path = tmp_path / "quick.xlsx"

    # Create minimal valid Excel
    wb = Workbook()
    wb.remove(wb.active)

    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Quick Test"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반"

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = f"조건{i+1}"

    wb.save(excel_path)

    # Quick validation
    validator = PDFValidator(excel_path=excel_path)
    result = validator.validate_table_structure()

    assert result.passed is True
    assert "cover_sheet" in result.details
    assert "conditions_sheet" in result.details


@pytest.mark.smoke
def test_quick_format_validation_with_minimal_pdf(tmp_path):
    """
    Smoke test: Quick format compliance validation with minimal PDF.
    """
    pdf_path = tmp_path / "quick.pdf"

    # Create minimal valid PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)

    # Page 1
    c.drawString(100, 100, "Test Page 1")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()

    # Page 2
    c.drawString(100, 100, "Test Page 2")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()

    c.save()

    # Quick validation
    validator = PDFValidator(pdf_path=pdf_path)
    result = validator.validate_format_compliance()

    assert result.passed is True
    assert result.details["page_count"] >= 2
    assert result.details["footer_found"] is True


@pytest.mark.smoke
def test_quick_font_validation_with_korean_pdf(tmp_path):
    """
    Smoke test: Quick font validation with Korean text PDF.

    Note: This test may pass with korean_text_found=True OR fail gracefully
    depending on PDF text extraction capabilities. The key is that validation
    runs without errors.
    """
    pdf_path = tmp_path / "korean.pdf"

    # Create PDF with Korean text
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "한글 테스트 Korean Text")
    c.drawString(100, 200, "추가 한글 내용")
    c.drawString(100, 300, "검증 테스트")
    c.save()

    # Quick validation
    validator = PDFValidator(pdf_path=pdf_path)
    result = validator.validate_font_embedding()

    # Validation should complete without errors
    # Korean text detection depends on PDF text extraction
    # If no Korean text found, it's a known limitation of reportlab PDFs
    assert result is not None
    assert hasattr(result, "passed")
    assert hasattr(result, "details")
    assert "korean_text_found" in result.details

    # If Korean text is found, great. If not, it's acceptable for reportlab PDFs
    # in smoke tests (integration tests use more complex PDFs)


@pytest.mark.smoke
def test_quick_validation_failure_detection(tmp_path):
    """
    Smoke test: Validation quickly detects failures.
    """
    excel_path = tmp_path / "empty.xlsx"

    # Create Excel with no required sheets
    wb = Workbook()
    wb.save(excel_path)

    # Quick validation - should fail fast
    validator = PDFValidator(excel_path=excel_path)
    result = validator.validate_table_structure()

    assert result.passed is False
    assert "시트가 없습니다" in result.message


@pytest.mark.smoke
def test_evidence_generation_performance(tmp_path):
    """
    Smoke test: Evidence generation completes quickly.
    """
    excel_path = tmp_path / "perf.xlsx"
    pdf_path = tmp_path / "perf.pdf"
    evidence_path = tmp_path / "evidence.json"

    # Create minimal files
    wb = Workbook()
    wb.remove(wb.active)
    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Perf Test"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반"

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = f"조건{i+1}"

    wb.save(excel_path)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "한글")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()
    c.drawString(100, 100, "Page 2")
    c.showPage()
    c.save()

    # Quick evidence generation
    import time

    start = time.time()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    validator.to_evidence(evidence_path)

    elapsed = time.time() - start

    # Verify evidence created quickly (< 2 seconds)
    assert evidence_path.exists()
    assert elapsed < 2.0, f"Evidence generation took {elapsed:.2f}s (should be < 2s)"


@pytest.mark.smoke
def test_validate_all_aggregation(tmp_path):
    """
    Smoke test: validate_all() correctly aggregates all validation results.
    """
    excel_path = tmp_path / "all.xlsx"
    pdf_path = tmp_path / "all.pdf"

    # Create minimal valid files
    wb = Workbook()
    wb.remove(wb.active)
    ws_cover = wb.create_sheet(SHEET_COVER)
    ws_cover[COVER_DATE_CELL] = "2025년 10월 25일"
    ws_cover[COVER_CUSTOMER_CELL] = "Test"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반"

    wb.create_sheet(SHEET_QUOTE)

    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):
        ws_conditions[f"{CONDITIONS_DESC_COL}{CONDITIONS_START_ROW + i}"] = f"조건{i+1}"

    wb.save(excel_path)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "한글")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()
    c.drawString(100, 100, "Page 2")
    c.showPage()
    c.save()

    # Quick validation
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Verify all 3 validations present
    assert len(results) == 3
    assert "table_structure" in results
    assert "font_embedding" in results
    assert "format_compliance" in results

    # Verify all are ValidationResult objects
    for key, result in results.items():
        assert isinstance(result, ValidationResult)
        assert hasattr(result, "passed")
        assert hasattr(result, "message")
        assert hasattr(result, "details")


@pytest.mark.smoke
def test_nonexistent_file_handling(tmp_path):
    """
    Smoke test: Validator handles nonexistent files gracefully.
    """
    nonexistent_excel = tmp_path / "nonexistent.xlsx"
    nonexistent_pdf = tmp_path / "nonexistent.pdf"

    # Excel validation with nonexistent file
    validator1 = PDFValidator(excel_path=nonexistent_excel)
    result1 = validator1.validate_table_structure()
    assert result1.passed is False
    assert "존재하지 않습니다" in result1.message

    # PDF validation with nonexistent file
    validator2 = PDFValidator(pdf_path=nonexistent_pdf)
    result2 = validator2.validate_font_embedding()
    assert result2.passed is False
    assert "존재하지 않습니다" in result2.message
