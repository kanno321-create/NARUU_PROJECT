"""Unit tests for Excel conditions sheet generation (_write_conditions_sheet)."""

import pytest
from openpyxl import Workbook
from kis_estimator_core.engine.excel_generator import ExcelGenerator
from kis_estimator_core.core.ssot.constants_format import (
    CONDITIONS_TITLE_CELL,
    CONDITIONS_START_ROW,
    CONDITIONS_ITEM_COL,
    CONDITIONS_DESC_COL,
    CONDITION_1_PANEL_NAME,
    CONDITION_2_ENCLOSURE_QTY,
    CONDITION_3_MAIN_QTY,
    CONDITION_4_BREAKER_SPEC,
    CONDITION_5_PANEL_SPEC,
    CONDITION_6_BRANCH_TYPES,
    CONDITION_7_REQUIRED_MATERIALS,
    CONDITION_8_BLANK_ROWS,
    CONDITION_9_매입함,
    CONDITION_10_MULTI_PANELS,
)


@pytest.fixture
def template_file(tmp_path):
    """Create a real Excel template file for testing."""
    template_path = tmp_path / "template.xlsx"
    wb = Workbook()
    wb.save(template_path)
    return template_path


@pytest.mark.unit
def test_write_conditions_sheet_title(template_file):
    """Test conditions sheet title is written correctly."""
    wb = Workbook()
    ws = wb.create_sheet("견적조건")

    # Create ExcelGenerator with real template
    gen = ExcelGenerator(template_path=template_file)

    # Call _write_conditions_sheet
    gen._write_conditions_sheet(ws)

    # Verify title
    assert ws[CONDITIONS_TITLE_CELL].value == "견적조건"


@pytest.mark.unit
def test_write_conditions_sheet_all_items(template_file):
    """Test all 10 conditions are written."""
    wb = Workbook()
    ws = wb.create_sheet("견적조건")

    gen = ExcelGenerator(template_path=template_file)
    gen._write_conditions_sheet(ws)

    # Verify all 10 items
    expected_conditions = [
        CONDITION_1_PANEL_NAME,
        CONDITION_2_ENCLOSURE_QTY,
        CONDITION_3_MAIN_QTY,
        CONDITION_4_BREAKER_SPEC,
        CONDITION_5_PANEL_SPEC,
        CONDITION_6_BRANCH_TYPES,
        CONDITION_7_REQUIRED_MATERIALS,
        CONDITION_8_BLANK_ROWS,
        CONDITION_9_매입함,
        CONDITION_10_MULTI_PANELS,
    ]

    for i, expected_text in enumerate(expected_conditions, start=1):
        row = CONDITIONS_START_ROW + (i - 1)

        # Check item number
        item_cell = f"{CONDITIONS_ITEM_COL}{row}"
        assert ws[item_cell].value == f"{i}."

        # Check description
        desc_cell = f"{CONDITIONS_DESC_COL}{row}"
        assert ws[desc_cell].value == expected_text


@pytest.mark.unit
def test_write_conditions_sheet_row_positions(template_file):
    """Test conditions are written at correct row positions."""
    wb = Workbook()
    ws = wb.create_sheet("견적조건")

    gen = ExcelGenerator(template_path=template_file)
    gen._write_conditions_sheet(ws)

    # Verify first condition at row 4
    assert ws[f"{CONDITIONS_ITEM_COL}{CONDITIONS_START_ROW}"].value == "1."

    # Verify last condition at row 13 (4 + 9)
    last_row = CONDITIONS_START_ROW + 9
    assert ws[f"{CONDITIONS_ITEM_COL}{last_row}"].value == "10."


@pytest.mark.unit
def test_write_conditions_sheet_idempotent(template_file):
    """Test writing conditions sheet twice produces same result."""
    wb = Workbook()
    ws = wb.create_sheet("견적조건")

    gen = ExcelGenerator(template_path=template_file)

    # Write twice
    gen._write_conditions_sheet(ws)
    gen._write_conditions_sheet(ws)

    # Verify still only 10 items (no duplication)
    assert ws[CONDITIONS_TITLE_CELL].value == "견적조건"
    assert ws[f"{CONDITIONS_ITEM_COL}{CONDITIONS_START_ROW}"].value == "1."
    assert ws[f"{CONDITIONS_ITEM_COL}{CONDITIONS_START_ROW + 9}"].value == "10."


@pytest.mark.unit
def test_write_conditions_sheet_korean_encoding(template_file):
    """Test Korean text is properly encoded in conditions."""
    wb = Workbook()
    ws = wb.create_sheet("견적조건")

    gen = ExcelGenerator(template_path=template_file)
    gen._write_conditions_sheet(ws)

    # Verify Korean characters are preserved
    # Check condition 9 (매입함) which has Korean text
    row_9 = CONDITIONS_START_ROW + 8  # 0-indexed offset
    desc_cell = f"{CONDITIONS_DESC_COL}{row_9}"

    assert "매입함" in ws[desc_cell].value
    assert isinstance(ws[desc_cell].value, str)
