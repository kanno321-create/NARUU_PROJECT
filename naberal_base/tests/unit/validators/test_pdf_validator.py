"""Unit tests for PDF validator (ValidationResult and PDFValidator)."""

import pytest
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from kis_estimator_core.validators.pdf_validator import ValidationResult, PDFValidator
from kis_estimator_core.core.ssot.constants_format import (
    SHEET_COVER,
    SHEET_QUOTE,
    SHEET_CONDITIONS,
    COVER_DATE_CELL,
    COVER_CUSTOMER_CELL,
    COVER_PROJECT_CELL,
)


@pytest.mark.unit
def test_validation_result_pass():
    """Test ValidationResult for passing validation."""
    result = ValidationResult(
        passed=True,
        message="All checks passed",
        details={"check1": "OK", "check2": "OK"},
    )

    assert result.passed is True
    assert "passed" in result.message.lower()
    assert result.details["check1"] == "OK"


@pytest.mark.unit
def test_validation_result_fail():
    """Test ValidationResult for failing validation."""
    result = ValidationResult(
        passed=False,
        message="Missing required sheet",
        details={"missing_sheet": "견적조건"},
    )

    assert result.passed is False
    assert "missing" in result.message.lower()


@pytest.mark.unit
def test_pdf_validator_init(tmp_path):
    """Test PDFValidator initialization."""
    # Create dummy files
    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"

    wb = Workbook()
    wb.save(excel_path)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.save()

    # Initialize validator
    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)

    assert validator.excel_path == excel_path
    assert validator.pdf_path == pdf_path
    assert validator.workbook is None  # Not loaded yet
    assert validator.pdf_reader is None  # Not loaded yet


@pytest.mark.unit
def test_validate_table_structure_missing_sheet(tmp_path):
    """Test table validation fails when required sheet is missing."""
    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"

    # Create Excel without required sheets
    wb = Workbook()
    wb.save(excel_path)

    # Create dummy PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.save()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    result = validator.validate_table_structure()

    # Should fail because no required sheets
    assert result.passed is False


@pytest.mark.unit
def test_validate_table_structure_all_sheets_present(tmp_path):
    """Test table validation passes when all sheets present."""
    from kis_estimator_core.core.ssot.constants_format import (
        COVER_PANEL_START_ROW,
        COVER_PANEL_NAME_COL,
        CONDITIONS_START_ROW,
        CONDITIONS_DESC_COL,
        CONDITIONS_TITLE_CELL,  # Added missing import
    )

    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"

    # Create Excel with all required sheets and cells
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    ws_cover = wb.create_sheet(SHEET_COVER)
    _ = wb.create_sheet(SHEET_QUOTE)
    ws_conditions = wb.create_sheet(SHEET_CONDITIONS)

    # Add required cells in cover sheet
    ws_cover[COVER_DATE_CELL] = "2025년 10월 24일"
    ws_cover[COVER_CUSTOMER_CELL] = "Test Customer"
    ws_cover[COVER_PROJECT_CELL] = "Test Project"
    ws_cover[f"{COVER_PANEL_NAME_COL}{COVER_PANEL_START_ROW}"] = "분전반1"  # Required!

    # Add conditions sheet title and items (minimum 5)
    ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"
    for i in range(5):  # Add 5 conditions minimum
        row = CONDITIONS_START_ROW + i
        ws_conditions[f"{CONDITIONS_DESC_COL}{row}"] = f"조건{i+1}"

    wb.save(excel_path)

    # Create dummy PDF
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.save()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    result = validator.validate_table_structure()

    assert result.passed is True
    assert "cover_sheet" in result.details
    assert "conditions_sheet" in result.details


@pytest.mark.unit
def test_validate_format_compliance_a4_size(tmp_path):
    """Test format validation checks A4 page size and required elements."""
    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"

    wb = Workbook()
    wb.save(excel_path)

    # Create PDF with A4 size, 2+ pages, and Footer markers
    c = canvas.Canvas(str(pdf_path), pagesize=A4)

    # Page 1
    c.drawString(100, 100, "Test Page 1")
    c.drawString(100, 50, "Build:test Hash:abc123 TS:2025-10-25")  # Footer markers
    c.showPage()

    # Page 2
    c.drawString(100, 100, "Test Page 2")
    c.drawString(100, 50, "Build:test Hash:abc123 TS:2025-10-25")  # Footer markers
    c.showPage()

    c.save()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    result = validator.validate_format_compliance()

    assert result.passed is True
    assert "page_count" in result.details
    assert result.details["page_count"] >= 2
    assert "footer_found" in result.details
    assert result.details["footer_found"] is True


@pytest.mark.unit
def test_validate_all_runs_all_checks(tmp_path):
    """Test validate_all() runs all three validation checks."""
    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"

    # Create minimal valid files
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(SHEET_COVER)
    wb.create_sheet(SHEET_QUOTE)
    wb.create_sheet(SHEET_CONDITIONS)
    wb.save(excel_path)

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "Test")
    c.save()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    results = validator.validate_all()

    # Should have 3 validation results
    assert "table_structure" in results
    assert "font_embedding" in results
    assert "format_compliance" in results

    # All should be ValidationResult objects
    assert isinstance(results["table_structure"], ValidationResult)
    assert isinstance(results["font_embedding"], ValidationResult)
    assert isinstance(results["format_compliance"], ValidationResult)


@pytest.mark.unit
def test_to_evidence_creates_json(tmp_path):
    """Test to_evidence() creates valid Evidence JSON."""
    excel_path = tmp_path / "test.xlsx"
    pdf_path = tmp_path / "test.pdf"
    evidence_path = tmp_path / "evidence.json"

    # Create minimal Excel file
    wb = Workbook()
    wb.save(excel_path)

    # Create 2-page PDF with footer markers
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(100, 100, "Page 1")
    c.drawString(100, 50, "Build:test Hash:abc TS:2025")
    c.showPage()
    c.drawString(100, 100, "Page 2")
    c.showPage()
    c.save()

    validator = PDFValidator(excel_path=excel_path, pdf_path=pdf_path)
    validator.validate_all()
    validator.to_evidence(evidence_path)

    # Verify JSON file created
    assert evidence_path.exists()

    # Verify JSON structure without reading (avoid encoding issues)
    import json

    # Read as binary, decode as UTF-8 explicitly
    with evidence_path.open("rb") as f:
        evidence_bytes = f.read()

    # Try to decode and parse
    try:
        evidence_text = evidence_bytes.decode("utf-8")
        evidence = json.loads(evidence_text)

        assert "timestamp" in evidence
        assert "results" in evidence
        assert "overall_passed" in evidence
    except UnicodeDecodeError:
        # If UTF-8 fails, try UTF-8-sig (BOM)
        evidence_text = evidence_bytes.decode("utf-8-sig")
        evidence = json.loads(evidence_text)

        assert "timestamp" in evidence
        assert "results" in evidence
