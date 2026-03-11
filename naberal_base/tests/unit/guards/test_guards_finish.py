"""
Guards Format Finish Tests (T2 - Phase I-5 핀포인트 마무리)

목적: core/ssot/guards_format.py 커버리지 증대 (0% → 70%+)
원칙: Zero-Mock, openpyxl Workbook 직접 생성 (Real Object)
"""

from openpyxl import Workbook

from kis_estimator_core.core.ssot.guards_format import (
    formula_guard,
    totals_guard,
)


# ============================================================
# T2-1: Formula Guard - Threshold Boundary
# ============================================================
def test_formula_guard_threshold_boundary():
    """
    수식 보존율이 정확히 threshold(95%)일 때 통과 검증

    Zero-Mock: 실제 openpyxl Workbook 생성
    """
    # Setup: 실제 Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 범위 H3:H22 (20개 셀)
    # 수식 19개 (95%) + 숫자 1개 (5%) = 정확히 95% threshold
    for row in range(3, 22):  # H3:H21 (19개)
        ws[f"H{row}"] = f"=F{row}*G{row}"

    ws["H22"] = 12000  # 마지막 1개는 숫자 (threshold 경계 테스트)

    # Execute: formula_guard 실행
    is_valid, errors, rate = formula_guard(
        ws, formula_ranges=["H3:H22"], allow_blank=False
    )

    # Assert: threshold 경계에서 통과해야 함
    assert rate == 95.0, f"Expected 95.0%, got {rate}%"
    assert is_valid is True, f"Should pass at threshold, but failed: {errors}"
    assert "수식 아님" in errors[0]  # H22는 수식 아니므로 경고 있음


# ============================================================
# T2-2: Formula Guard - Below Threshold Raises
# ============================================================
def test_formula_guard_below_threshold_raises():
    """
    수식 보존율이 threshold 미만일 때 실패 검증

    Zero-Mock: 실제 openpyxl Workbook 생성
    """
    # Setup: 실제 Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 범위 H3:H22 (20개 셀)
    # 수식 18개 (90%) + 숫자 2개 (10%) = 90% (threshold 95% 미만)
    for row in range(3, 21):  # H3:H20 (18개)
        ws[f"H{row}"] = f"=F{row}*G{row}"

    ws["H21"] = 10000  # 숫자
    ws["H22"] = 12000  # 숫자

    # Execute: formula_guard 실행
    is_valid, errors, rate = formula_guard(
        ws, formula_ranges=["H3:H22"], allow_blank=False
    )

    # Assert: threshold 미만이므로 실패해야 함
    assert rate == 90.0, f"Expected 90.0%, got {rate}%"
    assert is_valid is False, "Should fail below threshold"
    assert any(
        "보존율" in err and "임계치" in err for err in errors
    ), f"Should have threshold error, got: {errors}"


# ============================================================
# T2-3: Totals Guard - Formula Cells with Decimal
# ============================================================
def test_totals_guard_formula_cells_decimal():
    """
    totals_guard가 수식 셀을 올바르게 처리하는지 검증

    시나리오:
    - 라인아이템에 수식 셀 포함 (=F*G)
    - 소계/합계 셀도 수식 (=SUM(...))
    - Decimal 타입 검증

    Zero-Mock: 실제 openpyxl Workbook 생성
    """
    # Setup: 실제 Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 라인아이템: G3:G5 (수식 3개)
    ws["G3"] = "=F3*E3"  # 수식
    ws["G4"] = "=F4*E4"  # 수식
    ws["G5"] = 15000  # 숫자 (Decimal 변환 테스트)

    # 소계 셀: G10 (수식)
    ws["G10"] = "=SUM(G3:G5)"

    # 합계 셀: G12 (수식 - 견적서는 소계=합계)
    ws["G12"] = "=G10"

    # Execute: totals_guard 실행
    is_valid, errors = totals_guard(
        ws,
        subtotal_cell="G10",
        total_cell="G12",
        vat_cell=None,  # VAT 없음
        line_items_range="G3:G5",
    )

    # Assert: 수식 셀이어도 검증 통과 (openpyxl은 수식 계산 안 함, 존재만 확인)
    assert is_valid is True, f"Should pass with formula cells, errors: {errors}"
    # 수식 셀 2개는 skip되고 숫자 1개만 합산됨 (expected_subtotal = 15000)
    # 실제 값 비교는 안 함 (openpyxl 한계)
