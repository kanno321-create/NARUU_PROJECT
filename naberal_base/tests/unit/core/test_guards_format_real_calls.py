"""
guards_format.py 실제 호출 테스트 (P4-2 Real-Call 전환)

목적: coverage 측정을 위한 실제 함수 호출
원칙: Zero-Mock, SSOT 상수 사용, AppError 스키마 준수
"""

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from decimal import Decimal

from kis_estimator_core.core.ssot.guards_format import (
    named_range_guard,
    formula_guard,
    totals_guard,
    text_policy_guard,
    cell_protection_guard,
    nan_blank_guard,
    validate_date_format,
    validate_size_format,
    full_estimate_guard,
)
from kis_estimator_core.core.ssot.constants_format import (
    FORMULA_PRESERVE_THRESHOLD,
    TOTAL_TOLERANCE,
    SHEET_QUOTE,
    SHEET_COVER,
    NAMED_RANGE_QUOTE_ITEMS,
    REQUIRED_PHRASES,
    FORBIDDEN_PHRASES,
)


# ============================================================
# Helper: Create Test Workbook
# ============================================================
def _create_test_workbook_with_formulas():
    """테스트용 워크북 생성 (수식 포함)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 헤더
    ws["B1"] = "품목"
    ws["C1"] = "규격"
    ws["E1"] = "단위"
    ws["F1"] = "수량"
    ws["G1"] = "단가"
    ws["H1"] = "금액"

    # 데이터 행 (수식 포함)
    ws["B3"] = "외함"
    ws["C3"] = "600*800*200"
    ws["E3"] = "면"
    ws["F3"] = 1
    ws["G3"] = 100000
    ws["H3"] = "=G3*F3"  # 수식

    ws["B4"] = "차단기"
    ws["C4"] = "SBE-104"
    ws["E4"] = "EA"
    ws["F4"] = 1
    ws["G4"] = 50000
    ws["H4"] = "=G4*F4"  # 수식

    return wb


def _create_test_workbook_no_formulas():
    """테스트용 워크북 생성 (수식 없음 - threshold 미달)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 수식 없는 데이터
    ws["H3"] = 100000  # 숫자 (수식 아님)
    ws["H4"] = 50000  # 숫자 (수식 아님)

    return wb


# ============================================================
# Test: named_range_guard
# ============================================================
def test_named_range_guard_success():
    """네임드 범위 검증 성공 케이스"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 네임드 범위 정의
    named_range = DefinedName(
        NAMED_RANGE_QUOTE_ITEMS, attr_text=f"'{SHEET_QUOTE}'!$B$3:$H$10"
    )
    wb.defined_names.add(named_range)

    # 실제 호출
    is_valid, errors = named_range_guard(wb, required_ranges=[NAMED_RANGE_QUOTE_ITEMS])

    # 검증
    assert is_valid is True
    assert len(errors) == 0


def test_named_range_guard_missing_range():
    """네임드 범위 누락 케이스"""
    wb = Workbook()
    wb.active.title = SHEET_QUOTE

    # 네임드 범위 없음
    is_valid, errors = named_range_guard(wb, required_ranges=[NAMED_RANGE_QUOTE_ITEMS])

    # 검증: 실패
    assert is_valid is False
    assert len(errors) > 0
    assert any("누락" in err for err in errors)


# ============================================================
# Test: formula_guard
# ============================================================
def test_formula_guard_success():
    """수식 보존 검증 성공 케이스 (≥ THRESHOLD)"""
    wb = _create_test_workbook_with_formulas()
    ws = wb[SHEET_QUOTE]

    # 실제 호출
    is_valid, errors, preservation_rate = formula_guard(
        ws, formula_ranges=["H3:H4"], allow_blank=True
    )

    # 검증
    assert preservation_rate >= FORMULA_PRESERVE_THRESHOLD
    assert is_valid is True
    assert len(errors) == 0


def test_formula_guard_below_threshold():
    """수식 보존율 미달 케이스 (< THRESHOLD)"""
    wb = _create_test_workbook_no_formulas()
    ws = wb[SHEET_QUOTE]

    # 실제 호출
    is_valid, errors, preservation_rate = formula_guard(
        ws, formula_ranges=["H3:H4"], allow_blank=False  # strict mode
    )

    # 검증: 실패 (수식 없음)
    assert preservation_rate < FORMULA_PRESERVE_THRESHOLD
    assert is_valid is False
    assert len(errors) > 0


def test_formula_guard_error_formulas():
    """수식 오류 검증 (#REF!, #NAME!)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 오류 수식
    ws["H3"] = "=#REF!"
    ws["H4"] = "=#NAME?"

    # 실제 호출
    is_valid, errors, preservation_rate = formula_guard(
        ws, formula_ranges=["H3:H4"], allow_blank=False
    )

    # 검증: 실패 (오류 수식)
    assert is_valid is False
    assert any("#REF!" in err or "#NAME!" in err for err in errors)


# ============================================================
# Test: totals_guard
# ============================================================
def test_totals_guard_success():
    """소계/합계 일관성 검증 성공"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 라인아이템
    ws["G3"] = Decimal("100000.00")
    ws["G4"] = Decimal("50000.00")

    # 소계 (수식)
    ws["G48"] = "=SUM(G3:G4)"

    # 합계 (소계와 동일)
    ws["G50"] = "=G48"

    # 실제 호출
    is_valid, errors = totals_guard(
        ws,
        subtotal_cell="G48",
        total_cell="G50",
        vat_cell=None,
        tolerance=TOTAL_TOLERANCE,
        line_items_range="G3:G4",
    )

    # 검증
    assert is_valid is True
    assert len(errors) == 0


def test_totals_guard_missing_subtotal():
    """소계 셀 누락 케이스"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 소계 셀 없음
    # ws["G48"] = None (empty)
    ws["G50"] = 150000

    # 실제 호출
    is_valid, errors = totals_guard(
        ws, subtotal_cell="G48", total_cell="G50", vat_cell=None
    )

    # 검증: 실패
    assert is_valid is False
    assert any("소계 셀 누락" in err for err in errors)


# ============================================================
# Test: text_policy_guard
# ============================================================
def test_text_policy_guard_success():
    """텍스트 정책 검증 성공 (필수 문구 포함)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 필수 문구 포함 (금지 문자 제외)
    ws["B10"] = "납기 2주일"
    ws["B11"] = "대금 선불"
    ws["B12"] = "유효기간 30일"

    # 실제 호출
    is_valid, errors = text_policy_guard(
        ws,
        text_ranges=["B10:B12"],
        required_phrases=REQUIRED_PHRASES,
        forbidden_phrases=[],
    )

    # 검증
    assert is_valid is True
    assert len(errors) == 0


def test_text_policy_guard_forbidden_phrase():
    """금칙어 검증 실패"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 금칙어 포함
    ws["B10"] = "무료 배송"
    ws["B11"] = "서비스 제공"

    # 실제 호출
    is_valid, errors = text_policy_guard(
        ws,
        text_ranges=["B10:B11"],
        required_phrases=[],
        forbidden_phrases=FORBIDDEN_PHRASES,
    )

    # 검증: 실패
    assert is_valid is False
    assert len(errors) > 0
    assert any("금칙어" in err for err in errors)


# ============================================================
# Test: cell_protection_guard
# ============================================================
def test_cell_protection_guard_success():
    """셀 보호 검증 성공"""
    from openpyxl.styles import Protection

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 수식 셀 잠금 (보호) - Protection 객체는 immutable이므로 새로 생성
    ws["H3"].protection = Protection(locked=True)
    ws["H4"].protection = Protection(locked=True)

    # 입력 셀 잠금 해제
    ws["F3"].protection = Protection(locked=False)
    ws["G3"].protection = Protection(locked=False)

    # 실제 호출
    is_valid, errors = cell_protection_guard(
        ws, protected_ranges=["H3:H4"], unlocked_ranges=["F3", "G3"]
    )

    # 검증
    assert is_valid is True
    assert len(errors) == 0


def test_cell_protection_guard_unlocked_formula():
    """수식 셀 보호 실패 (잠금 해제됨)"""
    from openpyxl.styles import Protection

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 수식 셀이 잠금 해제됨 (보호 안 됨) - Protection 객체는 immutable이므로 새로 생성
    ws["H3"].protection = Protection(locked=False)

    # 실제 호출
    is_valid, errors = cell_protection_guard(
        ws, protected_ranges=["H3:H3"], unlocked_ranges=[]
    )

    # 검증: 실패
    assert is_valid is False
    assert any("미보호" in err for err in errors)


# ============================================================
# Test: nan_blank_guard
# ============================================================
def test_nan_blank_guard_success():
    """NaN/빈칸 검증 성공"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 정상 데이터
    ws["F3"] = 1
    ws["G3"] = 100000

    # 실제 호출
    is_valid, errors = nan_blank_guard(ws, critical_ranges=["F3:G3"], allow_zero=True)

    # 검증
    assert is_valid is True
    assert len(errors) == 0


def test_nan_blank_guard_blank_cell():
    """빈칸 검증 실패"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 빈칸
    # ws["F3"] = None (empty)
    ws["G3"] = 100000

    # 실제 호출
    is_valid, errors = nan_blank_guard(ws, critical_ranges=["F3:G3"], allow_zero=True)

    # 검증: 실패
    assert is_valid is False
    assert any("빈칸 발견" in err for err in errors)


# ============================================================
# Test: validate_date_format
# ============================================================
def test_validate_date_format_success():
    """날짜 형식 검증 성공"""
    # 실제 호출
    is_valid = validate_date_format("2025년 10월 17일")

    # 검증
    assert is_valid is True


def test_validate_date_format_failure():
    """날짜 형식 검증 실패"""
    # 실제 호출
    is_valid = validate_date_format("2025-10-17")  # 잘못된 형식

    # 검증
    assert is_valid is False


# ============================================================
# Test: validate_size_format
# ============================================================
def test_validate_size_format_success():
    """사이즈 형식 검증 성공"""
    # 실제 호출
    is_valid = validate_size_format("600*800*200")

    # 검증
    assert is_valid is True


def test_validate_size_format_failure():
    """사이즈 형식 검증 실패"""
    # 실제 호출
    is_valid = validate_size_format("600x800x200")  # 잘못된 구분자

    # 검증
    assert is_valid is False


# ============================================================
# Test: full_estimate_guard (Composite)
# ============================================================
def test_full_estimate_guard_minimal():
    """전체 검증 최소 케이스 (표지 없음)"""
    wb = _create_test_workbook_with_formulas()

    # 표지 시트 생성
    cover_sheet = wb.create_sheet(SHEET_COVER)
    cover_sheet["I17"] = "=H17*G17"

    # 실제 호출
    is_valid, results = full_estimate_guard(wb)

    # 검증: 결과 딕셔너리 확인
    assert "named_range" in results
    assert "formula_quote" in results
    assert "formula_cover" in results
    assert "totals" in results
    assert "cell_protection" in results
    assert "nan_blank" in results
