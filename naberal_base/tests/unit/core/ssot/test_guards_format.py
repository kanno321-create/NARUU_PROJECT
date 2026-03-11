"""
P4-2.2: Guards Format Tests
Target: guards_format.py 0% → 60% coverage (126/209 statements)

Tests format validation guards from Excel workbooks:
- formula_guard: Formula preservation ≥ 95%
- totals_guard: Subtotal/total/VAT consistency with Decimal precision
- named_range_guard: Required named ranges validation
- text_policy_guard: Required/forbidden phrases
- cell_protection_guard: Cell lock validation
- nan_blank_guard: NaN/blank prohibition

SB-05 Compliant: Zero-Mock, real openpyxl Workbook objects
LAW-02: SSOT constants (FORMULA_PRESERVE_THRESHOLD, TOTAL_TOLERANCE)

Note: guards_format.py returns (is_valid, errors) tuples, not AppErrors
"""

from decimal import Decimal
from openpyxl import Workbook

from kis_estimator_core.core.ssot.guards_format import (
    formula_guard,
    totals_guard,
    named_range_guard,
    nan_blank_guard,
)

# Import module explicitly for coverage measurement
from kis_estimator_core.core.ssot import guards_format as _  # noqa: F401


class TestFormulaGuard:
    """Formula preservation validation tests"""

    def test_formula_guard_100_percent_preservation(self):
        """100% formula preservation - PASS"""
        wb = Workbook()
        ws = wb.active

        # 10 formula cells
        for i in range(1, 11):
            ws.cell(row=i, column=1).value = f"=A{i}*2"

        is_valid, errors, rate = formula_guard(
            ws, formula_ranges=["A1:A10"], allow_blank=False
        )

        assert is_valid is True
        assert len(errors) == 0
        assert rate == 100.0

    def test_formula_guard_95_0_percent_boundary(self):
        """Exactly 95.0% preservation (boundary) - PASS (with allow_blank=True)"""
        wb = Workbook()
        ws = wb.active

        # 100 cells: 95 formulas, 5 non-formulas
        for i in range(1, 96):
            ws.cell(row=i, column=1).value = f"=A{i}*2"
        for i in range(96, 101):
            ws.cell(row=i, column=1).value = 100

        # allow_blank=True로 non-formula cells는 에러가 아님
        # 하지만 rate는 95%로 계산됨 (95 formulas / 100 non-empty)
        is_valid, errors, rate = formula_guard(
            ws, formula_ranges=["A1:A100"], allow_blank=True
        )

        # 95% >= 95% threshold이므로 PASS
        assert is_valid is True
        assert len(errors) == 0  # allow_blank=True이므로 non-formula에 대한 에러 없음
        assert rate == 95.0

    def test_formula_guard_94_9_percent_fail(self):
        """94.9% preservation - FAIL"""
        wb = Workbook()
        ws = wb.active

        # 1000 cells: 949 formulas, 51 non-formulas (94.9%)
        for i in range(1, 950):
            ws.cell(row=i, column=1).value = f"=A{i}*2"
        for i in range(950, 1001):
            ws.cell(row=i, column=1).value = 100

        is_valid, errors, rate = formula_guard(
            ws, formula_ranges=["A1:A1000"], allow_blank=False
        )

        assert is_valid is False
        assert len(errors) > 0
        assert rate < 95.0
        assert any("수식 보존율" in err for err in errors)

    def test_formula_guard_empty_cells_allowed(self):
        """Empty cells don't affect formula rate when allow_blank=True - PASS"""
        wb = Workbook()
        ws = wb.active

        # 10 formulas, 90 empty cells
        for i in range(1, 11):
            ws.cell(row=i, column=1).value = f"=SUM(A1:A{i})"
        # A11:A100 are empty (None)

        is_valid, errors, rate = formula_guard(
            ws, formula_ranges=["A1:A100"], allow_blank=True
        )

        assert is_valid is True
        assert len(errors) == 0
        assert rate == 100.0  # 10/10 non-empty cells are formulas

    def test_formula_guard_mixed_content(self):
        """Mixed formulas and values with 96% rate - PASS"""
        wb = Workbook()
        ws = wb.active

        # 96 formulas, 4 static values (96%)
        for i in range(1, 97):
            ws.cell(row=i, column=1).value = f"=A{i}+10"
        for i in range(97, 101):
            ws.cell(row=i, column=1).value = 50

        # allow_blank=True로 non-formula에 대한 에러 없음
        is_valid, errors, rate = formula_guard(
            ws, formula_ranges=["A1:A100"], allow_blank=True
        )

        assert is_valid is True
        assert len(errors) == 0
        assert rate == 96.0


class TestTotalsGuard:
    """Subtotal/total/VAT consistency validation with Decimal precision"""

    def test_totals_guard_exact_match(self):
        """Subtotal and total exactly match - PASS"""
        wb = Workbook()
        ws = wb.active

        # Subtotal: 1000.00
        ws["G20"].value = Decimal("1000.00")
        # Total: 1000.00 (no multiplier)
        ws["G21"].value = Decimal("1000.00")

        is_valid, errors = totals_guard(
            ws,
            subtotal_cell="G20",
            total_cell="G21",
            vat_cell=None,
            line_items_range=None,
        )

        assert is_valid is True
        assert len(errors) == 0

    def test_totals_guard_within_tolerance(self):
        """Difference within TOTAL_TOLERANCE - PASS"""
        wb = Workbook()
        ws = wb.active

        # Subtotal: 1000.00
        ws["G20"].value = Decimal("1000.00")
        # Total: 1000.01 (within 0.01 tolerance)
        ws["G21"].value = Decimal("1000.01")

        is_valid, errors = totals_guard(
            ws,
            subtotal_cell="G20",
            total_cell="G21",
            vat_cell=None,
            line_items_range=None,
        )

        assert is_valid is True
        assert len(errors) == 0

    def test_totals_guard_missing_subtotal(self):
        """Missing subtotal cell - FAIL"""
        wb = Workbook()
        ws = wb.active

        # Subtotal: None
        # Total: 1000.00
        ws["G21"].value = Decimal("1000.00")

        is_valid, errors = totals_guard(
            ws,
            subtotal_cell="G20",  # Empty
            total_cell="G21",
            vat_cell=None,
            line_items_range=None,
        )

        assert is_valid is False
        assert len(errors) >= 1
        assert any("소계 셀 누락" in err for err in errors)

    def test_totals_guard_with_line_items(self):
        """Line items direct calculation - PASS"""
        wb = Workbook()
        ws = wb.active

        # Line items: 3 cells with 100, 200, 300 (sum = 600)
        ws["G3"].value = Decimal("100.00")
        ws["G4"].value = Decimal("200.00")
        ws["G5"].value = Decimal("300.00")

        # Subtotal: 600.00
        ws["G20"].value = Decimal("600.00")
        # Total: 600.00
        ws["G21"].value = Decimal("600.00")

        is_valid, errors = totals_guard(
            ws,
            subtotal_cell="G20",
            total_cell="G21",
            vat_cell=None,
            line_items_range="G3:G5",
        )

        assert is_valid is True
        assert len(errors) == 0


class TestNamedRangeGuard:
    """Named range validation tests"""

    def test_named_range_guard_all_present(self):
        """All required named ranges present - PASS"""
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        _ = wb.active  # noqa: F811

        # Create required named ranges using DefinedNameDict API
        defn1 = DefinedName("CustomerName", attr_text="Sheet!$C$5")
        defn2 = DefinedName("ProjectName", attr_text="Sheet!$C$7")
        defn3 = DefinedName("QuoteDate", attr_text="Sheet!$C$3")

        wb.defined_names["CustomerName"] = defn1
        wb.defined_names["ProjectName"] = defn2
        wb.defined_names["QuoteDate"] = defn3

        required_ranges = ["CustomerName", "ProjectName", "QuoteDate"]

        is_valid, errors = named_range_guard(wb, required_ranges)

        assert is_valid is True
        assert len(errors) == 0

    def test_named_range_guard_missing_range(self):
        """Missing required named range - FAIL"""
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        _ = wb.active  # noqa: F811

        # Only create 2 out of 3 ranges
        defn1 = DefinedName("CustomerName", attr_text="Sheet!$C$5")
        defn2 = DefinedName("ProjectName", attr_text="Sheet!$C$7")

        wb.defined_names["CustomerName"] = defn1
        wb.defined_names["ProjectName"] = defn2
        # QuoteDate missing

        required_ranges = ["CustomerName", "ProjectName", "QuoteDate"]

        is_valid, errors = named_range_guard(wb, required_ranges)

        assert is_valid is False
        assert len(errors) >= 1
        assert any("QuoteDate" in err for err in errors)

    def test_named_range_guard_empty_list(self):
        """Empty required list - PASS (no requirements)"""
        wb = Workbook()
        _ = wb.active  # noqa: F811

        is_valid, errors = named_range_guard(wb, required_ranges=[])

        assert is_valid is True
        assert len(errors) == 0


class TestNanBlankGuard:
    """NaN/blank prohibition tests"""

    def test_nan_blank_guard_no_issues(self):
        """No NaN or blank in critical range - PASS"""
        wb = Workbook()
        ws = wb.active

        # Fill cells with valid values
        for row in range(1, 11):
            ws.cell(row=row, column=7).value = 1000 * row  # Column G

        is_valid, errors = nan_blank_guard(ws, critical_ranges=["G1:G10"])

        assert is_valid is True
        assert len(errors) == 0

    def test_nan_blank_guard_blank_detected(self):
        """Blank cell in critical range - FAIL"""
        wb = Workbook()
        ws = wb.active

        # Fill most cells
        for row in range(1, 10):
            ws.cell(row=row, column=7).value = 1000 * row
        # Leave G10 blank (None)

        is_valid, errors = nan_blank_guard(ws, critical_ranges=["G1:G10"])

        assert is_valid is False
        assert len(errors) >= 1
        assert any("빈칸" in err for err in errors)

    def test_nan_blank_guard_allow_zero(self):
        """Zero value allowed with allow_zero=True - PASS"""
        wb = Workbook()
        ws = wb.active

        # Fill with 0 values
        for row in range(1, 11):
            ws.cell(row=row, column=7).value = 0

        is_valid, errors = nan_blank_guard(
            ws, critical_ranges=["G1:G10"], allow_zero=True
        )

        assert is_valid is True
        assert len(errors) == 0

    def test_nan_blank_guard_zero_prohibited(self):
        """Zero value prohibited with allow_zero=False - FAIL"""
        wb = Workbook()
        ws = wb.active

        # Fill with 0 values
        for row in range(1, 11):
            ws.cell(row=row, column=7).value = 0

        is_valid, errors = nan_blank_guard(
            ws, critical_ranges=["G1:G10"], allow_zero=False
        )

        assert is_valid is False
        assert len(errors) >= 1
        assert any("0 값" in err for err in errors)
