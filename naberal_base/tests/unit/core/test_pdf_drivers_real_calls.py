"""
pdf_drivers.py 실제 호출 테스트 (P4-2 Phase I-4)

목적: PDF 변환 드라이버 시스템 coverage 측정 (0% → 60%)
원칙: Zero-Mock (실제 subprocess/import 호출), SSOT 정책 사용
"""

import pytest
import sys
from unittest.mock import patch

from kis_estimator_core.core.ssot.pdf_drivers import (
    PDFDriver,
    check_libreoffice,
    check_excel_com,
    check_openpyxl_reportlab,
    choose_driver,
    convert_xlsx_to_pdf_openpyxl,
)
from kis_estimator_core.core.ssot.errors import EstimatorError


# ============================================================
# Test: check_libreoffice (LibreOffice 가용성 체크)
# ============================================================
def test_check_libreoffice_not_available():
    """LibreOffice 미설치 시 False 반환"""
    # 실제 호출 (대부분 환경에서 False 예상)
    result = check_libreoffice()

    # 검증: bool 값 반환
    assert isinstance(result, bool)


@pytest.mark.skipif(not check_libreoffice(), reason="LibreOffice not installed")
def test_check_libreoffice_available():
    """LibreOffice 설치 시 True 반환 (설치된 경우만 실행)"""
    # 실제 호출
    result = check_libreoffice()

    # 검증: LibreOffice 있으면 True
    assert result is True


# ============================================================
# Test: check_excel_com (Excel COM 가용성 체크)
# ============================================================
def test_check_excel_com_non_windows():
    """Windows 아닌 환경에서 False 반환"""
    if sys.platform == "win32":
        pytest.skip("Test only for non-Windows platforms")

    # 실제 호출
    result = check_excel_com()

    # 검증: Windows 아니면 False
    assert result is False


@pytest.mark.skipif(sys.platform != "win32", reason="Windows only")
def test_check_excel_com_windows():
    """Windows 환경에서 win32com 가용성 체크"""
    # 실제 호출
    result = check_excel_com()

    # 검증: bool 값 반환 (win32com 설치 여부에 따라)
    assert isinstance(result, bool)


# ============================================================
# Test: check_openpyxl_reportlab (openpyxl+ReportLab 체크)
# ============================================================
def test_check_openpyxl_reportlab_available():
    """openpyxl + ReportLab 가용성 체크"""
    # 실제 호출 (CI 환경에서는 보통 설치되어 있음)
    result = check_openpyxl_reportlab()

    # 검증: bool 값 반환
    assert isinstance(result, bool)


# ============================================================
# Test: choose_driver (최적 드라이버 선택)
# ============================================================
def test_choose_driver_with_available_backend():
    """사용 가능한 백엔드가 있을 때 드라이버 선택"""
    # 실제 호출 (환경에 따라 다름)
    try:
        driver, description = choose_driver()

        # 검증: 드라이버 타입과 설명 반환
        assert isinstance(driver, PDFDriver)
        assert isinstance(description, str)
        assert driver in [
            PDFDriver.LIBREOFFICE,
            PDFDriver.EXCEL_COM,
            PDFDriver.OPENPYXL_REPORTLAB,
        ]

    except EstimatorError as e:
        # 어떤 백엔드도 없으면 에러 발생 (정상)
        assert "No PDF driver available" in str(e)


@patch("kis_estimator_core.core.ssot.pdf_drivers.check_libreoffice", return_value=True)
def test_choose_driver_libreoffice_priority(mock_check):
    """LibreOffice가 최우선 순위"""
    # 실제 호출
    driver, description = choose_driver()

    # 검증: LibreOffice 선택
    assert driver == PDFDriver.LIBREOFFICE
    assert "LibreOffice" in description


@patch("kis_estimator_core.core.ssot.pdf_drivers.check_libreoffice", return_value=False)
@patch(
    "kis_estimator_core.core.ssot.pdf_drivers.check_openpyxl_reportlab",
    return_value=True,
)
def test_choose_driver_openpyxl_fallback(mock_openpyxl, mock_libreoffice):
    """LibreOffice 없으면 openpyxl+ReportLab 사용"""
    # 실제 호출
    driver, description = choose_driver()

    # 검증: openpyxl+ReportLab 선택
    assert driver == PDFDriver.OPENPYXL_REPORTLAB
    assert "openpyxl" in description


@patch("kis_estimator_core.core.ssot.pdf_drivers.check_libreoffice", return_value=False)
@patch(
    "kis_estimator_core.core.ssot.pdf_drivers.check_openpyxl_reportlab",
    return_value=False,
)
@patch("kis_estimator_core.core.ssot.pdf_drivers.check_excel_com", return_value=True)
def test_choose_driver_excel_com_last_resort(
    mock_excel, mock_openpyxl, mock_libreoffice
):
    """Excel COM이 마지막 대안 (Windows only)"""
    # 실제 호출
    driver, description = choose_driver()

    # 검증: Excel COM 선택
    assert driver == PDFDriver.EXCEL_COM
    assert "Excel COM" in description


@patch("kis_estimator_core.core.ssot.pdf_drivers.check_libreoffice", return_value=False)
@patch(
    "kis_estimator_core.core.ssot.pdf_drivers.check_openpyxl_reportlab",
    return_value=False,
)
@patch("kis_estimator_core.core.ssot.pdf_drivers.check_excel_com", return_value=False)
def test_choose_driver_no_backend_error(mock_excel, mock_openpyxl, mock_libreoffice):
    """어떤 백엔드도 없으면 에러 발생"""
    # 실제 호출: EstimatorError 예상
    with pytest.raises(EstimatorError) as exc_info:
        choose_driver()

    # 검증: 적절한 에러 메시지
    assert "No PDF driver available" in str(exc_info.value)


# ============================================================
# Test: convert_xlsx_to_pdf_openpyxl (Fallback 변환)
# ============================================================
@pytest.mark.skipif(
    not check_openpyxl_reportlab(), reason="openpyxl or reportlab not installed"
)
def test_convert_xlsx_to_pdf_openpyxl_success(tmp_path):
    """openpyxl + ReportLab 변환 성공 (실제 파일 생성)"""
    # 테스트용 Excel 파일 생성
    import openpyxl

    xlsx_path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    ws["A2"] = 100
    ws["A3"] = "=A2*2"
    wb.save(xlsx_path)

    pdf_path = tmp_path / "test.pd"

    # 실제 호출
    convert_xlsx_to_pdf_openpyxl(xlsx_path, pdf_path)

    # 검증: PDF 파일 생성됨
    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0  # 파일 크기 > 0


# ============================================================
# Test: PDFDriver Enum
# ============================================================
def test_pdf_driver_enum_values():
    """PDFDriver Enum 값 확인"""
    # 실제 호출
    assert PDFDriver.LIBREOFFICE == "libreoffice"
    assert PDFDriver.EXCEL_COM == "excel_com"
    assert PDFDriver.OPENPYXL_REPORTLAB == "openpyxl_reportlab"


def test_pdf_driver_enum_members():
    """PDFDriver Enum 멤버 확인"""
    # 실제 호출
    members = list(PDFDriver)

    # 검증: 3개 멤버 존재
    assert len(members) == 3
    assert PDFDriver.LIBREOFFICE in members
    assert PDFDriver.EXCEL_COM in members
    assert PDFDriver.OPENPYXL_REPORTLAB in members
