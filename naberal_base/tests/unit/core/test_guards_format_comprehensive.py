"""
guards_format.py COMPREHENSIVE TESTS - Missing Branch Coverage
P4-2 Coverage Expansion - Target: All uncovered branches in guards_format.py
"""

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from decimal import Decimal

from kis_estimator_core.core.ssot.guards_format import (
    named_range_guard,
    formula_guard,
    totals_guard,
    text_policy_guard,
    nan_blank_guard,
    cell_protection_guard,
)
from kis_estimator_core.core.ssot.constants_format import (
    SHEET_QUOTE,
    REQUIRED_PHRASES,
)


# ===========================================================
# Test: named_range_guard - Missing Branches
# ===========================================================
def test_named_range_guard_none_value():
    """네임드 범위가 None인 케이스 (line 142)"""
    wb = Workbook()
    wb.active.title = SHEET_QUOTE

    # 네임드 범위 생성하되 value=None
    named_range = DefinedName("test_range", attr_text=None)
    wb.defined_names.add(named_range)

    # 실제 호출
    is_valid, errors = named_range_guard(wb, required_ranges=["test_range"])

    # 검증: None value → 손상 오류
    assert is_valid is False
    assert any("손상" in err for err in errors)


# ===========================================================
# Test: formula_guard - policy_on_empty Branches
# ===========================================================
def test_formula_guard_strict_fail_on_empty():
    """policy_on_empty='strict_fail' - 빈 셀 즉시 FAIL (line 191-193)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 빈 셀 포함 범위
    ws["H3"] = "=G3*F3"  # 수식
    # H4 is empty
    ws["H5"] = "=G5*F5"  # 수식

    # 실제 호출
    is_valid, errors, rate = formula_guard(
        ws, formula_ranges=["H3:H5"], policy_on_empty="strict_fail"
    )

    # 검증: strict_fail → 빈 셀 발견 시 에러
    assert is_valid is False
    assert any("빈 셀 발견 (strict)" in err for err in errors)


def test_formula_guard_allow_blank_false_non_formula():
    """allow_blank=False + 수식 아님 → 에러 (line 231-232)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 수식이 아닌 값
    ws["H3"] = 100000  # 숫자 (수식 아님)
    ws["H4"] = "text"  # 텍스트 (수식 아님)

    # 실제 호출
    is_valid, errors, rate = formula_guard(
        ws, formula_ranges=["H3:H4"], allow_blank=False
    )

    # 검증: 수식 아님 → 에러
    assert is_valid is False
    assert any("수식 아님" in err for err in errors)


# ===========================================================
# Test: totals_guard - Decimal Conversion Errors
# ===========================================================
def test_totals_guard_invalid_type():
    """타입 위반 - 숫자가 아닌 값 (line 322-326)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 잘못된 타입
    ws["G3"] = "not a number"  # 문자열
    ws["G4"] = Decimal("50000.00")

    ws["G48"] = "=SUM(G3:G4)"
    ws["G50"] = "=G48"

    # 실제 호출
    is_valid, errors = totals_guard(
        ws,
        subtotal_cell="G48",
        total_cell="G50",
        vat_cell=None,
        line_items_range="G3:G4",
    )

    # 검증: 타입 위반 에러
    assert is_valid is False
    assert any("타입 위반" in err for err in errors)


def test_totals_guard_decimal_conversion_fail():
    """Decimal 변환 실패 - 잘못된 문자열 (line 333-336)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # Decimal 변환 불가능한 문자열
    ws["G3"] = "invalid_number"  # 숫자 아님
    ws["G4"] = Decimal("50000.00")

    ws["G48"] = "=SUM(G3:G4)"
    ws["G50"] = "=G48"

    # 실제 호출
    is_valid, errors = totals_guard(
        ws,
        subtotal_cell="G48",
        total_cell="G50",
        vat_cell=None,
        line_items_range="G3:G4",
    )

    # 검증: 타입 위반 에러
    assert is_valid is False
    assert any("타입 위반" in err for err in errors)


def test_totals_guard_missing_total():
    """합계 셀 누락 (line 354-356)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    ws["G48"] = Decimal("150000.00")
    # G50 is empty (합계 누락)

    # 실제 호출
    is_valid, errors = totals_guard(
        ws, subtotal_cell="G48", total_cell="G50", vat_cell=None
    )

    # 검증: 합계 셀 누락 에러
    assert is_valid is False
    assert any("합계 셀 누락" in err for err in errors)


# ===========================================================
# Test: text_policy_guard - Forbidden Chars
# ===========================================================
def test_text_policy_guard_forbidden_chars():
    """금지 문자 발견 (line 428-430)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 금지 문자 포함
    ws["B10"] = "파일명: <test>.txt"  # < > 금지
    ws["B11"] = "경로: C:\\folder\\file"  # \\ 금지

    # 실제 호출
    is_valid, errors = text_policy_guard(
        ws, text_ranges=["B10:B11"], required_phrases=[], forbidden_phrases=[]
    )

    # 검증: 금지 문자 발견 에러
    assert is_valid is False
    assert any("금지 문자" in err for err in errors)


def test_text_policy_guard_missing_required():
    """필수 문구 누락 (line 433-435)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 필수 문구 누락 (REQUIRED_PHRASES 중 일부만 포함)
    ws["B10"] = "납기: 2주일"
    # "대금", "유효기간" 등 누락

    # 실제 호출
    is_valid, errors = text_policy_guard(
        ws,
        text_ranges=["B10:B10"],
        required_phrases=REQUIRED_PHRASES,
        forbidden_phrases=[],
    )

    # 검증: 필수 문구 누락 에러
    assert is_valid is False
    assert any("필수 문구 누락" in err for err in errors)


# ===========================================================
# Test: nan_blank_guard - NaN and Zero Detection
# ===========================================================
def test_nan_blank_guard_nan_detection():
    """NaN 발견 (line 539-542)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # NaN 값
    ws["F3"] = float("nan")
    ws["G3"] = 100000

    # 실제 호출
    is_valid, errors = nan_blank_guard(ws, critical_ranges=["F3:G3"], allow_zero=True)

    # 검증: NaN 발견 에러
    assert is_valid is False
    assert any("NaN 발견" in err for err in errors)


def test_nan_blank_guard_zero_not_allowed():
    """0 값 발견 (allow_zero=False) (line 535-536)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 0 값
    ws["F3"] = 0
    ws["G3"] = 100000

    # 실제 호출
    is_valid, errors = nan_blank_guard(
        ws, critical_ranges=["F3:G3"], allow_zero=False  # 0 불허
    )

    # 검증: 0 값 발견 에러
    assert is_valid is False
    assert any("0 값 발견" in err for err in errors)


# ===========================================================
# Test: cell_protection_guard - Locked Input Cell
# ===========================================================
def test_cell_protection_guard_locked_input():
    """입력 셀이 잠김 (line 487-492)"""
    from openpyxl.styles import Protection

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 입력 셀이 잠김 (보호됨) - openpyxl Protection은 immutable이므로 새 객체 할당
    ws["F3"].protection = Protection(locked=True, hidden=False)  # 입력 셀인데 잠김

    # 실제 호출
    is_valid, errors = cell_protection_guard(
        ws, protected_ranges=[], unlocked_ranges=["F3"]  # F3은 잠금 해제되어야 함
    )

    # 검증: 입력 셀 잠김 에러
    assert is_valid is False
    assert any("입력 셀 잠금됨" in err for err in errors)


# ===========================================================
# Test: Validate Size Format - Alternative Formats
# ===========================================================
def test_validate_size_format_with_multiply_sign():
    """사이즈 형식 검증 - × 기호 사용"""
    from kis_estimator_core.core.ssot.guards_format import validate_size_format

    # × 기호 사용 (W×H×D)
    is_valid = validate_size_format("600×800×200")

    # 검증: × 기호도 허용
    assert is_valid is True


# ===========================================================
# Test: Formula Guard - Blank Cell with allow_blank=True
# ===========================================================
def test_formula_guard_blank_allowed_skip():
    """allow_blank=True - 빈 셀 스킵 (line 217-218)"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_QUOTE

    # 빈 셀 포함 범위
    ws["H3"] = "=G3*F3"  # 수식
    # H4 is empty (skipped)
    ws["H5"] = "=G5*F5"  # 수식

    # 실제 호출
    is_valid, errors, rate = formula_guard(
        ws, formula_ranges=["H3:H5"], allow_blank=True  # 빈 셀 허용
    )

    # 검증: 빈 셀 스킵, 수식 셀만 계산
    assert is_valid is True
    assert rate == 100.0  # 2/2 = 100%
