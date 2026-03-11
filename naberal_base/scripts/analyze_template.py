"""
템플릿 구조 분석 스크립트
견적서양식.xlsx 파일의 수식, 네임드 범위, 셀 구조 분석

실행: python scripts/analyze_template.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
from pathlib import Path


def analyze_template(template_path: str):
    """템플릿 파일 구조 분석"""

    print(f"[Template Analysis] Starting: {template_path}\n")

    # 1. 워크북 로드
    wb = openpyxl.load_workbook(template_path, data_only=False)

    print(f"[OK] Workbook loaded")
    print(f"    Sheets: {wb.sheetnames}\n")

    analysis = {
        "sheets": {},
        "named_ranges": {},
        "formulas": [],
        "data_areas": {},
    }

    # 2. 네임드 범위 분석
    print("[Named Ranges]:")
    if wb.defined_names:
        for name, defn in wb.defined_names.items():
            analysis["named_ranges"][name] = str(defn.attr_text)
            print(f"   - {name}: {defn.attr_text}")
    else:
        print("   (네임드 범위 없음)")
    print()

    # 3. 각 시트 분석
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[Sheet] {sheet_name}")

        sheet_info = {
            "dimensions": str(ws.dimensions),
            "formulas": [],
            "merged_cells": [],
            "data_structure": [],
        }

        # 병합 셀
        if ws.merged_cells:
            sheet_info["merged_cells"] = [str(mc) for mc in ws.merged_cells.ranges]
            print(f"    Merged cells: {len(sheet_info['merged_cells'])}")

        # 수식 찾기
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    formula_info = {
                        "cell": cell.coordinate,
                        "formula": cell.value,
                    }
                    sheet_info["formulas"].append(formula_info)
                    analysis["formulas"].append({
                        "sheet": sheet_name,
                        **formula_info
                    })

        print(f"    Formulas: {formula_count}")

        # 데이터 영역 탐지 (첫 10행 샘플)
        print(f"    Data structure (first 10 rows):")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                if cell.value:
                    col_letter = get_column_letter(col_idx)
                    row_data.append({
                        "cell": f"{col_letter}{row_idx}",
                        "value": str(cell.value)[:50],  # 최대 50자
                        "type": type(cell.value).__name__,
                    })
            if row_data:
                sheet_info["data_structure"].append({
                    "row": row_idx,
                    "cells": row_data,
                })
                print(f"      Row {row_idx}: {len(row_data)} cells")

        analysis["sheets"][sheet_name] = sheet_info
        print()

    # 4. 결과 저장
    output_path = Path("claudedocs/template_analysis.json")
    output_path.parent.mkdir(exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)

    print(f"[OK] Analysis complete!")
    print(f"    Result saved: {output_path}")
    print(f"\n[Summary]:")
    print(f"    - Sheets: {len(analysis['sheets'])}")
    print(f"    - Named ranges: {len(analysis['named_ranges'])}")
    print(f"    - Total formulas: {len(analysis['formulas'])}")

    return analysis


if __name__ == "__main__":
    template_path = r"C:\Users\PC\Desktop\NABERAL_PROJECT-master\절대코어파일\견적서양식.xlsx"
    analyze_template(template_path)
