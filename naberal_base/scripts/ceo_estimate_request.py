#!/usr/bin/env python3
"""
대표님 요청 견적서 생성
- 외함: 옥외 SUS 201 1.2T
- 메인: LS 4P 250A (MCCB ABN-204)
- 분기: ELB 4P 100A ×2, ELB 4P 50A ×4, ELB 2P 20A ×12
"""

import sys
import asyncio
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.kis_estimator_core.engine.enclosure_solver import EnclosureSolver
from src.kis_estimator_core.models.enclosure import (
    BreakerSpec,
    CustomerRequirements,
)

# ==============================================================================
# 견적 사양 (대표님 요청)
# ==============================================================================
ESTIMATE_SPEC = {
    "customer_name": "대표님",
    "project_name": "CEO 요청 견적",
    "enclosure": {
        "type": "옥외노출",
        "material": "SUS201",
        "thickness": "1.2T",
    },
    "main_breaker": {
        "brand": "LS산전",
        "type": "MCCB",
        "model": "ABN-204",
        "poles": 4,
        "current_a": 250,
        "frame_af": 200,  # 200AF
        "breaking_capacity_ka": 26,  # CEO 정정: 26kA (2024-12-02)
        "price": 154000,  # CEO 지시로 추가된 가격
        "dimensions_mm": (140, 165, 60),  # W×H×D
    },
    "branch_breakers": [
        # ELB 4P 100A × 2 (경제형 EBN-104)
        {
            "brand": "LS산전",
            "type": "ELB",
            "model": "EBN-104",
            "poles": 4,
            "current_a": 100,
            "frame_af": 100,
            "breaking_capacity_ka": 18,  # CEO 정정: 18kA (2024-12-02)
            "price": 104170,  # CEO 정정: 104,170원 (2024-12-02)
            "quantity": 2,
            "dimensions_mm": (100, 130, 60),  # 4P 100AF 경제형
        },
        # ELB 4P 50A × 4 (표준형 EBS-54 - 4P 50AF 경제형 없음)
        {
            "brand": "LS산전",
            "type": "ELB",
            "model": "EBS-54",
            "poles": 4,
            "current_a": 50,
            "frame_af": 50,
            "breaking_capacity_ka": 18,  # CEO 정정: 18kA (2024-12-02)
            "price": 87340,  # CEO 정정: 87,340원 (2024-12-02)
            "quantity": 4,
            "dimensions_mm": (100, 130, 60),  # 4P 50AF
        },
        # ELB 2P 20A × 12 (소형 32GRHS)
        {
            "brand": "LS산전",
            "type": "ELB",
            "model": "32GRHS",
            "poles": 2,
            "current_a": 20,
            "frame_af": 30,  # 소형
            "breaking_capacity_ka": 2.5,
            "price": 10300,  # CEO 정정 가격
            "quantity": 12,
            "dimensions_mm": (33, 70, 42),  # 소형
        },
    ],
}


async def calculate_enclosure_size():
    """외함 크기 계산 (Stage 1)"""
    solver = EnclosureSolver()

    # 메인차단기
    main_spec = ESTIMATE_SPEC["main_breaker"]
    main_breaker = BreakerSpec(
        id="MAIN",
        model=main_spec["model"],
        poles=main_spec["poles"],
        current_a=main_spec["current_a"],
        frame_af=main_spec["frame_af"],
    )

    # 분기차단기
    branch_breakers = []
    for branch_spec in ESTIMATE_SPEC["branch_breakers"]:
        for i in range(branch_spec["quantity"]):
            branch_breakers.append(
                BreakerSpec(
                    id=f"{branch_spec['model']}_{i+1}",
                    model=branch_spec["model"],
                    poles=branch_spec["poles"],
                    current_a=branch_spec["current_a"],
                    frame_af=branch_spec["frame_af"],
                )
            )

    # 고객 요구사항
    enc_spec = ESTIMATE_SPEC["enclosure"]
    customer_reqs = CustomerRequirements(
        enclosure_type=enc_spec["type"],
        material=enc_spec["material"],
        ip_rating="IP44",
    )

    # 외함 계산 (async)
    result = await solver.solve(
        main_breaker=main_breaker,
        branch_breakers=branch_breakers,
        accessories=[],
        customer_requirements=customer_reqs,
    )

    return result


def get_enclosure_price(width_mm: int, height_mm: int, depth_mm: int) -> int:
    """
    SC 시리즈 (옥외 SUS201 1.2T) 기성함 가격 조회
    CEO 지시: 기성함 카탈로그에서 존재 여부 확인 필수
    SC 시리즈 표준 높이: 600, 700, 800, 900, 1000, 1200 (1100 없음!)
    """
    # SC 시리즈 가격표 (CEO 학습 규칙 반영)
    sc_prices = {
        # (W, H, D): price
        (700, 600, 200): 170000,
        (700, 700, 200): 190000,
        (700, 800, 200): 218000,
        (700, 900, 200): 246000,
        (700, 1000, 200): 266000,
        (700, 1200, 200): 320000,
        (800, 600, 200): 190000,
        (800, 700, 200): 210000,
        (800, 800, 200): 240000,
        (800, 900, 200): 270000,
        (800, 1000, 200): 295000,
        (800, 1200, 200): 355000,
        # 추가 규격
        (600, 600, 200): 150000,
        (600, 700, 200): 165000,
        (600, 800, 200): 185000,
        (600, 900, 200): 210000,
        (600, 1000, 200): 235000,
        (600, 1200, 200): 285000,
    }

    # 표준 높이 목록 (CEO 지시: 1100mm 없음)
    standard_heights = [600, 700, 800, 900, 1000, 1200]

    # 계산된 높이보다 크거나 같은 가장 작은 표준 높이 선택
    selected_height = None
    for h in standard_heights:
        if h >= height_mm:
            selected_height = h
            break
    if selected_height is None:
        selected_height = 1200  # 최대값

    # 깊이 표준화 (150, 200, 250)
    if depth_mm <= 150:
        selected_depth = 150
    elif depth_mm <= 200:
        selected_depth = 200
    else:
        selected_depth = 250

    # 실제 깊이는 기본 200mm 사용 (SC 시리즈 기본)
    selected_depth = 200

    # 폭 표준화
    if width_mm <= 600:
        selected_width = 600
    elif width_mm <= 700:
        selected_width = 700
    else:
        selected_width = 800

    key = (selected_width, selected_height, selected_depth)
    price = sc_prices.get(key)

    if price is None:
        # 가격이 없으면 면적 기반 계산
        area = selected_width * selected_height
        price = int((area / 90000) * 32000) + 100000

    return price, (selected_width, selected_height, selected_depth)


def calculate_accessories(total_breakers: int, main_frame_af: int, enclosure_h: int):
    """부속자재 계산"""
    accessories = []

    # 1. E.T (Earth Terminal)
    # 수량 공식: 12개당 1개 추가
    et_qty = (total_breakers // 12) + 1
    # 단가: 메인 AF 기준
    if main_frame_af <= 250:
        et_price = 4500
    elif main_frame_af == 400:
        et_price = 12000
    else:
        et_price = 18000
    accessories.append({"name": "E.T", "spec": "", "unit": "EA", "qty": et_qty, "price": et_price})

    # 2. N.T (Neutral Terminal)
    accessories.append({"name": "N.T", "spec": "", "unit": "EA", "qty": 1, "price": 3000})

    # 3. N.P (Name Plate)
    # 분기차단기 수량만큼 CARD HOLDER
    branch_count = total_breakers - 1  # 메인 제외
    accessories.append({"name": "N.P", "spec": "CARD HOLDER", "unit": "EA", "qty": branch_count, "price": 800})
    accessories.append({"name": "N.P", "spec": "3T×40×200", "unit": "EA", "qty": 1, "price": 4000})

    # 4. MAIN BUS-BAR
    # 규격: 메인 AF 기준 (CEO 규칙)
    if main_frame_af <= 125:
        main_busbar_spec = "3T×15"
    elif main_frame_af <= 250:
        main_busbar_spec = "5T×20"
    elif main_frame_af == 400:
        main_busbar_spec = "6T×30"
    else:
        main_busbar_spec = "8T×40"
    # 중량 계산 (면적 × 계수)
    if main_frame_af <= 250:
        coef = 0.000007
    elif main_frame_af <= 400:
        coef = 0.000013
    else:
        coef = 0.000015
    busbar_weight = round(700 * enclosure_h * coef, 1)
    # CEO 정정: 부스바 단가 22,000원/kg (2024-12-02)
    accessories.append({"name": "MAIN BUS-BAR", "spec": main_busbar_spec, "unit": "KG", "qty": busbar_weight, "price": 22000})

    # 5. BUS-BAR (분기용)
    # CEO 규칙: 분기부스바 스펙은 분기차단기 중 최대 AF 기준 (현재 100AF → 3T×15)
    branch_busbar_spec = "3T×15"  # 분기 최대 100AF 기준
    branch_busbar_weight = round(busbar_weight * 0.6, 1)
    # CEO 정정: 부스바 단가 22,000원/kg (2024-12-02)
    accessories.append({"name": "BUS-BAR", "spec": branch_busbar_spec, "unit": "KG", "qty": branch_busbar_weight, "price": 22000})

    # 6. COATING
    accessories.append({"name": "COATING", "spec": "PVC(20mm)", "unit": "M", "qty": 1, "price": 5000})

    # 7. P-COVER (아크릴)
    # 단가: ((W×H) ÷ 90,000) × 3,200원
    pcover_price = int((700 * enclosure_h / 90000) * 3200)
    accessories.append({"name": "P-COVER", "spec": "아크릴(PC)", "unit": "EA", "qty": 1, "price": pcover_price})

    # 8. ELB지지대 (소형 누전차단기용)
    # 32GRHS 12개
    accessories.append({"name": "ELB지지대", "spec": "", "unit": "EA", "qty": 12, "price": 500})

    # 9. INSULATOR
    accessories.append({"name": "INSULATOR", "spec": "EPOXY 40×40", "unit": "EA", "qty": 4, "price": 1100})

    # 10. 잡자재비
    # 기본 7,000원 + H 100mm당 1,000원, 최대 40,000원
    misc_price = min(7000 + (enclosure_h // 100) * 1000, 40000)
    accessories.append({"name": "잡자재비", "spec": "", "unit": "식", "qty": 1, "price": misc_price})

    # 11. ASSEMBLY CHARGE (조립비)
    # CEO 규칙 (2024-12-02): 메인차단기 AF 기준으로 차단기당 단가 결정
    # 100AF까지: 2,000원, 125AF~250AF: 3,000원, 400AF: 5,000원, 600~800AF: 6,000원
    if main_frame_af <= 100:
        assembly_per_breaker = 2000
    elif main_frame_af <= 250:
        assembly_per_breaker = 3000
    elif main_frame_af <= 400:
        assembly_per_breaker = 5000
    else:  # 600~800AF
        assembly_per_breaker = 6000
    assembly_price = 50000 + total_breakers * assembly_per_breaker
    accessories.append({"name": "ASSEMBLY CHARGE", "spec": "", "unit": "식", "qty": 1, "price": assembly_price})

    return accessories


async def create_estimate_excel():
    """견적서 Excel 생성"""
    print("=" * 80)
    print("대표님 요청 견적서 생성")
    print("=" * 80)

    # 사양 출력
    print("\n[견적 사양]")
    enc = ESTIMATE_SPEC["enclosure"]
    print(f"- 외함: {enc['type']} {enc['material']} {enc['thickness']}")
    main = ESTIMATE_SPEC["main_breaker"]
    print(f"- 메인: {main['brand']} {main['type']} {main['model']} {main['poles']}P {main['current_a']}A")
    print("- 분기:")
    for branch in ESTIMATE_SPEC["branch_breakers"]:
        print(f"  * {branch['type']} {branch['poles']}P {branch['current_a']}A × {branch['quantity']}개 ({branch['model']})")

    # Stage 1: 외함 크기 계산
    print("\n[Stage 1] 외함 크기 계산...")
    enclosure_result = await calculate_enclosure_size()
    dims = enclosure_result.dimensions
    print(f"  계산된 크기: {dims.width_mm}×{dims.height_mm}×{dims.depth_mm} mm")

    # 기성함 규격 조회 (CEO 지시: 반드시 카탈로그 확인)
    enc_price, (enc_w, enc_h, enc_d) = get_enclosure_price(dims.width_mm, dims.height_mm, dims.depth_mm)
    print(f"  기성함 선택: {enc_w}×{enc_h}×{enc_d} mm (SC 시리즈)")
    print(f"  외함 가격: {enc_price:,}원")

    # 총 차단기 수
    total_breakers = 1  # 메인
    for branch in ESTIMATE_SPEC["branch_breakers"]:
        total_breakers += branch["quantity"]
    print(f"\n[정보] 총 차단기 수: {total_breakers}개")

    # 부속자재 계산
    main_spec = ESTIMATE_SPEC["main_breaker"]
    accessories = calculate_accessories(total_breakers, main_spec["frame_af"], enc_h)

    # Excel 파일 생성
    wb = Workbook()

    # ===== 견적서 시트 =====
    ws = wb.active
    ws.title = "견적서"

    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # 헤더
    ws["A1"] = " 공종명 : "
    ws["B1"] = "분전반"
    ws["B3"] = f"{enc['type']} {enc['material']} {enc['thickness']}"

    ws["A2"] = "번호"
    ws["B2"] = "품명"
    ws["C2"] = "규격"
    ws["D2"] = "단위"
    ws["E2"] = "수량"
    ws["F2"] = "단   가"
    ws["G2"] = "금   액"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f"{col}2"].font = header_font
        ws[f"{col}2"].alignment = center_align

    row = 3
    item_num = 1

    # 1. 외함
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = f"{enc['type']} {enc['material']} {enc['thickness']}"
    ws[f"C{row}"] = f"{enc_w}×{enc_h}×{enc_d}"
    ws[f"D{row}"] = "면"
    ws[f"E{row}"] = 1
    ws[f"F{row}"] = enc_price
    ws[f"G{row}"] = enc_price
    row += 1
    item_num += 1

    # 2. 메인차단기
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = f"{main_spec['type']} ({main_spec['model']})"
    ws[f"C{row}"] = f"{main_spec['poles']}P {main_spec['frame_af']}AF {main_spec['current_a']}AT {main_spec['breaking_capacity_ka']}kA"
    ws[f"D{row}"] = "EA"
    ws[f"E{row}"] = 1
    ws[f"F{row}"] = main_spec["price"]
    ws[f"G{row}"] = main_spec["price"]
    row += 1
    item_num += 1

    # 3. 분기차단기
    for branch in ESTIMATE_SPEC["branch_breakers"]:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = f"{branch['type']} ({branch['model']})"
        ws[f"C{row}"] = f"{branch['poles']}P {branch['frame_af']}AF {branch['current_a']}AT {branch['breaking_capacity_ka']}kA"
        ws[f"D{row}"] = "EA"
        ws[f"E{row}"] = branch["quantity"]
        ws[f"F{row}"] = branch["price"]
        ws[f"G{row}"] = branch["price"] * branch["quantity"]
        row += 1
        item_num += 1

    # 4. 부속자재
    for acc in accessories:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = acc["name"]
        ws[f"C{row}"] = acc["spec"]
        ws[f"D{row}"] = acc["unit"]
        ws[f"E{row}"] = acc["qty"]
        ws[f"F{row}"] = acc["price"]
        ws[f"G{row}"] = int(acc["price"] * acc["qty"])
        row += 1
        item_num += 1

    # 빈 행 (최소 5열 공백)
    start_empty = row
    for _ in range(5):
        ws[f"G{row}"] = 0
        row += 1

    # 소계
    ws[f"B{row}"] = "소     계"
    ws[f"D{row}"] = "식"
    ws[f"E{row}"] = 1
    ws[f"G{row}"] = f"=SUM(G3:G{row-1})"
    subtotal_row = row
    row += 1

    # 합계
    ws[f"B{row}"] = "합     계"
    ws[f"D{row}"] = "식"
    ws[f"E{row}"] = 1
    ws[f"G{row}"] = f"=G{subtotal_row}"
    total_row = row

    # ===== 표지 시트 =====
    ws_cover = wb.create_sheet("표지", 0)

    # 표지 헤더
    ws_cover.merge_cells("A1:K1")
    ws_cover["A1"] = "견   적   서"
    ws_cover["A1"].font = Font(size=24, bold=True)
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_cover["A3"] = "일       자: "
    ws_cover["C3"] = datetime.now().strftime("%Y년 %m월 %d일")

    ws_cover["G3"] = "㈜ 한 국 산 업"
    ws_cover["G5"] = "주소:   경기도 안산시 단원구 시화벤처로 467"
    ws_cover["G6"] = "TEL  031-431-1413"
    ws_cover["G7"] = "FAX  031-431-1419"
    ws_cover["G10"] = "대표이사 :  李   忠   源   (인)"

    ws_cover["A5"] = "수       신:           "
    ws_cover["C5"] = ESTIMATE_SPEC["customer_name"]

    ws_cover["A7"] = "건       명:           "
    ws_cover["C7"] = ESTIMATE_SPEC["project_name"]

    ws_cover["A11"] = "신한은행 100-031-963940"

    # 합계
    ws_cover["A15"] = " 합   계 : "

    # 품목 헤더
    ws_cover["A16"] = "순서"
    ws_cover["B16"] = "품    명"
    ws_cover["D16"] = "규    격(가로×세로×깊이)"
    ws_cover["F16"] = "단위"
    ws_cover["G16"] = "수량"
    ws_cover["H16"] = "단   가"
    ws_cover["I16"] = "금   액"
    ws_cover["J16"] = "비   고"

    # 분전반 품목
    ws_cover["A17"] = 1
    ws_cover["B17"] = "분전반"
    ws_cover["D17"] = f"{enc_w}×{enc_h}×{enc_d}"
    ws_cover["F17"] = "면"
    ws_cover["G17"] = 1
    ws_cover["H17"] = f"=+견적서!G{total_row}"
    ws_cover["I17"] = "=H17*G17"
    ws_cover["J17"] = enc['type']

    # 합계 행
    ws_cover["B18"] = "     합         계"
    ws_cover["F18"] = "식"
    ws_cover["G18"] = "=SUM(G17:G17)"
    ws_cover["I18"] = "=SUM(I17:I17)"

    # 부가세
    ws_cover["H19"] = "부가세포함"
    ws_cover["I19"] = "=I18*1.1"

    # NOTE
    ws_cover["A20"] = f"< NOTE > 1. 차단기: {main_spec['brand']}, 외함: {enc['type']} {enc['material']} {enc['thickness']} (SC 시리즈)"
    ws_cover["A21"] = "             2. 제작도면 및 승인도면 제작비 별도"
    ws_cover["A22"] = "             3. 운임비 별도 (화물배송, 택배배송, 1톤 용달)"
    ws_cover["A24"] = "결제조건:200만원이하 출고전 완불/ 200만원이상 계약금30% 출고전 완불"

    # 파일 저장
    output_dir = project_root / "outputs"
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_dir / f"CEO_estimate_{timestamp}.xlsx"
    wb.save(filename)

    print("\n" + "=" * 80)
    print("견적서 생성 완료!")
    print("=" * 80)
    print(f"파일: {filename}")
    print(f"외함: {enc_w}×{enc_h}×{enc_d} mm ({enc['type']} {enc['material']} {enc['thickness']})")
    print(f"메인: {main_spec['model']} {main_spec['poles']}P {main_spec['current_a']}A ({main_spec['price']:,}원)")
    print(f"분기: ELB 4P 100A × 2, ELB 4P 50A × 4, ELB 2P 20A × 12")
    print("=" * 80)

    return filename


if __name__ == "__main__":
    output_file = asyncio.run(create_estimate_excel())
    print(f"\n[완료] {output_file}")

    # Excel 파일 열기
    import subprocess
    subprocess.run(["start", str(output_file)], shell=True, check=False)
