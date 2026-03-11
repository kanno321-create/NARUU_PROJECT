"""
ERP 거래처/상품 일괄 등록 스크립트
- 거래처: 한국산업_거래처등록.xlsx → Railway API POST /v1/erp/customers
- 상품: 한국산업_상품리스트.xlsx → Railway API POST /v1/erp/products
- 도매가 필드 제외

사용법:
    python scripts/bulk_import_erp.py --target customers
    python scripts/bulk_import_erp.py --target products
    python scripts/bulk_import_erp.py --target all
"""

import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl
import requests

# ============================================================================
# 설정
# ============================================================================

RAILWAY_URL = "https://naberalproject-production.up.railway.app"
LOGIN_URL = f"{RAILWAY_URL}/v1/auth/login"
CUSTOMER_URL = f"{RAILWAY_URL}/v1/erp/customers"
PRODUCT_URL = f"{RAILWAY_URL}/v1/erp/products"

# 로그인 정보
USERNAME = "ceo"
PASSWORD = "ceo1234"

# Excel 파일 경로
BASE_DIR = Path(__file__).parent.parent / "절대코어파일"
CUSTOMER_FILE = BASE_DIR / "한국산업_거래처등록.xlsx"
PRODUCT_FILE = BASE_DIR / "한국산업_상품리스트.xlsx"

# 배치 설정
BATCH_DELAY = 0.05  # 요청 간 50ms 대기 (rate limit 방지)


# ============================================================================
# 인증
# ============================================================================

def get_auth_token() -> str:
    """Railway API 로그인 후 JWT 토큰 반환"""
    print(f"[AUTH] {RAILWAY_URL} 로그인 중...")
    resp = requests.post(
        LOGIN_URL,
        json={"username": USERNAME, "password": PASSWORD},
        timeout=30,
    )
    if resp.status_code != 200:
        print(f"[AUTH] 로그인 실패: {resp.status_code} {resp.text}")
        sys.exit(1)

    token = resp.json()["access_token"]
    print(f"[AUTH] 로그인 성공")
    return token


def auth_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


# ============================================================================
# 거래처 등록
# ============================================================================

def parse_customers() -> list[dict]:
    """Excel → 거래처 데이터 파싱"""
    wb = openpyxl.load_workbook(str(CUSTOMER_FILE), data_only=True)
    ws = wb[wb.sheetnames[0]]

    customers = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value
        if not name or not str(name).strip():
            continue

        cat = ws.cell(row=row_idx, column=3).value or "매출처"
        # 거래처구분 매핑: 매입매출처 → 매출매입 (ERP 스키마에 맞춤)
        cat_str = str(cat).strip()
        if cat_str == "매입매출처":
            cat_str = "매출매입"

        phone = ws.cell(row=row_idx, column=4).value or ""
        mobile = ws.cell(row=row_idx, column=5).value or ""
        fax = ws.cell(row=row_idx, column=6).value or ""
        email = ws.cell(row=row_idx, column=7).value or ""
        zipcode = ws.cell(row=row_idx, column=8).value or ""
        addr1 = ws.cell(row=row_idx, column=9).value or ""
        addr2 = ws.cell(row=row_idx, column=10).value or ""
        ceo = ws.cell(row=row_idx, column=11).value or ""
        biz_num = ws.cell(row=row_idx, column=12).value or ""

        # 주소 합치기
        address_parts = []
        if str(zipcode).strip():
            address_parts.append(f"[{str(zipcode).strip()}]")
        if str(addr1).strip():
            address_parts.append(str(addr1).strip())
        if str(addr2).strip():
            address_parts.append(str(addr2).strip())
        address = " ".join(address_parts)

        # 전화번호: 전화 없으면 핸드폰 사용
        phone_str = str(phone).strip() if phone else ""
        if not phone_str and mobile:
            phone_str = str(mobile).strip()

        # 핸드폰 번호를 메모에 보존 (전화번호와 다를 때)
        memo_parts = []
        if mobile and str(mobile).strip() and str(mobile).strip() != phone_str:
            memo_parts.append(f"핸드폰: {str(mobile).strip()}")

        customers.append({
            "name": str(name).strip(),
            "customer_type": cat_str,
            "business_number": str(biz_num).strip() if biz_num else None,
            "ceo_name": str(ceo).strip() if ceo else None,
            "phone": phone_str or None,
            "fax": str(fax).strip() if fax else None,
            "email": str(email).strip() if email else None,
            "address": address or None,
            "memo": "; ".join(memo_parts) if memo_parts else None,
            "grade": "일반",
            "is_active": True,
            "credit_limit": 0,
            "current_receivable": 0,
        })

    wb.close()
    return customers


def import_customers(token: str, dry_run: bool = False):
    """거래처 일괄 등록"""
    customers = parse_customers()
    print(f"\n[거래처] 총 {len(customers)}건 파싱 완료")

    if dry_run:
        print("[거래처] DRY RUN — 샘플 3건:")
        for c in customers[:3]:
            print(f"  {json.dumps(c, ensure_ascii=False, indent=2)}")
        return

    headers = auth_headers(token)
    success = 0
    fail = 0
    errors = []

    for i, cust in enumerate(customers):
        try:
            resp = requests.post(
                CUSTOMER_URL, json=cust, headers=headers, timeout=30
            )
            if resp.status_code in (200, 201):
                success += 1
            else:
                fail += 1
                errors.append(f"행{i+2} {cust['name']}: {resp.status_code} {resp.text[:100]}")
        except Exception as e:
            fail += 1
            errors.append(f"행{i+2} {cust['name']}: {str(e)}")

        # 진행률 표시 (50건마다)
        if (i + 1) % 50 == 0:
            print(f"  [{i+1}/{len(customers)}] 성공: {success}, 실패: {fail}")

        time.sleep(BATCH_DELAY)

    print(f"\n[거래처] 완료: 성공 {success}건, 실패 {fail}건")
    if errors:
        print(f"[거래처] 오류 목록 (처음 10건):")
        for e in errors[:10]:
            print(f"  - {e}")


# ============================================================================
# 상품 등록
# ============================================================================

def parse_products() -> list[dict]:
    """Excel → 상품 데이터 파싱 (도매가 제외)"""
    wb = openpyxl.load_workbook(str(PRODUCT_FILE), data_only=True)
    ws = wb[wb.sheetnames[0]]

    products = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name or not str(name).strip():
            continue

        spec = ws.cell(row=row_idx, column=2).value or ""
        # 상세규격(col3)은 99.8% 비어있으므로 무시
        code = ws.cell(row=row_idx, column=4).value or ""
        retail = ws.cell(row=row_idx, column=5).value or 0  # 소매가 → selling_price
        # col6 도매가 — 제외
        purchase = ws.cell(row=row_idx, column=7).value or 0  # 입고가 → purchase_price
        memo = ws.cell(row=row_idx, column=8).value or ""

        # 코드 정리
        code_str = str(code).strip()
        if code_str == "(제작)":
            code_str = None  # 자동 생성되도록

        # 가격 정리
        try:
            sell_price = float(retail)
        except (ValueError, TypeError):
            sell_price = 0
        try:
            buy_price = float(purchase)
        except (ValueError, TypeError):
            buy_price = 0

        products.append({
            "name": str(name).strip(),
            "code": code_str,
            "spec": str(spec).strip() if spec else None,
            "unit": "EA",
            "selling_price": sell_price,
            "purchase_price": buy_price,
            "memo": str(memo).strip() if memo else None,
            "safety_stock": 0,
            "is_active": True,
        })

    wb.close()
    return products


def import_products(token: str, dry_run: bool = False):
    """상품 일괄 등록 (도매가 제외)"""
    products = parse_products()
    print(f"\n[상품] 총 {len(products)}건 파싱 완료 (도매가 제외)")

    if dry_run:
        print("[상품] DRY RUN — 샘플 5건:")
        for p in products[:5]:
            print(f"  {json.dumps(p, ensure_ascii=False, indent=2)}")
        return

    headers = auth_headers(token)
    success = 0
    fail = 0
    errors = []

    for i, prod in enumerate(products):
        try:
            resp = requests.post(
                PRODUCT_URL, json=prod, headers=headers, timeout=30
            )
            if resp.status_code in (200, 201):
                success += 1
            else:
                fail += 1
                errors.append(f"행{i+2} {prod['name']}: {resp.status_code} {resp.text[:100]}")
        except Exception as e:
            fail += 1
            errors.append(f"행{i+2} {prod['name']}: {str(e)}")

        # 진행률 표시 (100건마다)
        if (i + 1) % 100 == 0:
            print(f"  [{i+1}/{len(products)}] 성공: {success}, 실패: {fail}")

        time.sleep(BATCH_DELAY)

    print(f"\n[상품] 완료: 성공 {success}건, 실패 {fail}건")
    if errors:
        print(f"[상품] 오류 목록 (처음 10건):")
        for e in errors[:10]:
            print(f"  - {e}")


# ============================================================================
# 메인
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="ERP 거래처/상품 일괄 등록")
    parser.add_argument(
        "--target",
        choices=["customers", "products", "all"],
        default="all",
        help="등록 대상 (customers, products, all)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="실제 등록 없이 파싱 결과만 확인",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("ERP 일괄 등록 스크립트")
    print(f"대상: {args.target}")
    print(f"모드: {'DRY RUN (시뮬레이션)' if args.dry_run else '실제 등록'}")
    print("=" * 60)

    token = None
    if not args.dry_run:
        token = get_auth_token()

    if args.target in ("customers", "all"):
        import_customers(token, dry_run=args.dry_run)

    if args.target in ("products", "all"):
        import_products(token, dry_run=args.dry_run)

    print("\n" + "=" * 60)
    print("완료")
    print("=" * 60)


if __name__ == "__main__":
    main()