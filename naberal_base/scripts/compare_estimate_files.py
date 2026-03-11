"""
Excel Template vs Real Estimate Comparison Tool
Compares template structure with actual filled estimate to understand data mapping.

CRITICAL: NO MOCK DATA - Only analyze real files
"""

import openpyxl
import json
from pathlib import Path
from typing import Dict, List, Any


def analyze_sheet_detailed(sheet) -> Dict[str, Any]:
    """Analyze sheet structure with detailed cell-by-cell data."""
    result = {
        "name": sheet.title,
        "dimensions": str(sheet.dimensions),
        "row_count": sheet.max_row,
        "col_count": sheet.max_column,
        "merged_cells": [str(r) for r in sheet.merged_cells.ranges],
        "formulas": [],
        "data_rows": []
    }

    # Collect all formulas
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                result["formulas"].append({
                    "cell": cell.coordinate,
                    "formula": cell.value
                })

    # Collect data rows (first 60 rows for detailed analysis)
    for row_idx, row in enumerate(sheet.iter_rows(max_row=60, values_only=False), start=1):
        row_data = {
            "row": row_idx,
            "cells": []
        }

        for cell in row:
            if cell.value is not None:
                cell_info = {
                    "col": cell.column_letter,
                    "coord": cell.coordinate,
                    "value": str(cell.value)[:100] if cell.value else None,  # Truncate long values
                    "type": "formula" if (isinstance(cell.value, str) and cell.value.startswith('=')) else "data"
                }

                # Add number format if it's a number
                if isinstance(cell.value, (int, float)):
                    cell_info["number_format"] = cell.number_format

                row_data["cells"].append(cell_info)

        if row_data["cells"]:  # Only add non-empty rows
            result["data_rows"].append(row_data)

    return result


def compare_files(template_path: str, example_path: str) -> Dict[str, Any]:
    """Compare template and real example files."""
    print("[Template Analysis]")
    print(f"Template: {template_path}")
    print(f"Example:  {example_path}")

    # Load both files
    wb_template = openpyxl.load_workbook(template_path, data_only=False)
    wb_example = openpyxl.load_workbook(example_path, data_only=False)

    comparison = {
        "template": {
            "file": template_path,
            "sheet_count": len(wb_template.sheetnames),
            "sheet_names": wb_template.sheetnames,
            "sheets": {}
        },
        "example": {
            "file": example_path,
            "sheet_count": len(wb_example.sheetnames),
            "sheet_names": wb_example.sheetnames,
            "sheets": {}
        },
        "comparison": {
            "sheet_name_match": wb_template.sheetnames == wb_example.sheetnames,
            "differences": []
        }
    }

    # Analyze template sheets
    print("\n[Analyzing Template Sheets]")
    for sheet_name in wb_template.sheetnames:
        print(f"  - {sheet_name}")
        comparison["template"]["sheets"][sheet_name] = analyze_sheet_detailed(wb_template[sheet_name])

    # Analyze example sheets
    print("\n[Analyzing Example Sheets]")
    for sheet_name in wb_example.sheetnames:
        print(f"  - {sheet_name}")
        comparison["example"]["sheets"][sheet_name] = analyze_sheet_detailed(wb_example[sheet_name])

    # Compare structures
    print("\n[Comparing Structures]")

    # Compare sheet names
    template_sheets = set(wb_template.sheetnames)
    example_sheets = set(wb_example.sheetnames)

    if template_sheets != example_sheets:
        comparison["comparison"]["differences"].append({
            "type": "sheet_names",
            "template_only": list(template_sheets - example_sheets),
            "example_only": list(example_sheets - template_sheets)
        })

    # Compare common sheets
    common_sheets = template_sheets & example_sheets
    for sheet_name in common_sheets:
        t_sheet = comparison["template"]["sheets"][sheet_name]
        e_sheet = comparison["example"]["sheets"][sheet_name]

        diff = {
            "sheet": sheet_name,
            "dimension_match": t_sheet["dimensions"] == e_sheet["dimensions"],
            "merged_cells_match": t_sheet["merged_cells"] == e_sheet["merged_cells"],
            "formula_count": {
                "template": len(t_sheet["formulas"]),
                "example": len(e_sheet["formulas"])
            }
        }

        comparison["comparison"]["differences"].append(diff)

    return comparison


def find_data_mapping(comparison: Dict[str, Any]) -> Dict[str, Any]:
    """Identify data mapping patterns between template and example."""
    print("\n[Finding Data Mapping Patterns]")

    mapping = {
        "cover_sheet": {},
        "estimate_sheet": {},
        "patterns": []
    }

    # Analyze 견적서 (estimate details) sheet
    estimate_name = "견적서"
    if estimate_name in comparison["template"]["sheets"] and estimate_name in comparison["example"]["sheets"]:
        t_sheet = comparison["template"]["sheets"][estimate_name]
        e_sheet = comparison["example"]["sheets"][estimate_name]

        print(f"\n[{estimate_name} Sheet Analysis]")
        print(f"Template rows: {t_sheet['row_count']}, Example rows: {e_sheet['row_count']}")

        # Find data rows (where template has formulas but example has data)
        estimate_mapping = {
            "header_row": None,
            "data_start_row": None,
            "data_end_row": None,
            "column_mapping": {},
            "filled_rows": []
        }

        # Find header row (Row 2 based on previous analysis)
        if len(e_sheet["data_rows"]) >= 2:
            header_row = e_sheet["data_rows"][1]  # Row 2 (index 1)
            estimate_mapping["header_row"] = 2

            # Extract column headers
            for cell in header_row["cells"]:
                estimate_mapping["column_mapping"][cell["col"]] = cell["value"]

        # Find data rows (Row 3 onwards)
        data_rows = []
        for row_data in e_sheet["data_rows"][2:]:  # Skip first 2 rows
            if row_data["row"] >= 3:
                # Check if row has actual data (not just formulas)
                has_data = any(
                    cell["type"] == "data"
                    for cell in row_data["cells"]
                    if cell["col"] in ["A", "B", "C", "D", "E", "F"]
                )

                if has_data:
                    data_rows.append({
                        "row": row_data["row"],
                        "data": {cell["col"]: cell["value"] for cell in row_data["cells"]}
                    })

        estimate_mapping["data_start_row"] = 3
        estimate_mapping["data_end_row"] = data_rows[-1]["row"] if data_rows else None
        estimate_mapping["filled_rows"] = data_rows
        estimate_mapping["total_items"] = len(data_rows)

        mapping["estimate_sheet"] = estimate_mapping

        print(f"Data range: Row {estimate_mapping['data_start_row']} to {estimate_mapping['data_end_row']}")
        print(f"Total items: {estimate_mapping['total_items']}")

    # Analyze 표지 (cover) sheet
    cover_name = "표지"
    if cover_name in comparison["example"]["sheets"]:
        e_sheet = comparison["example"]["sheets"][cover_name]

        cover_mapping = {
            "customer_info": {},
            "project_info": {},
            "summary_table": {}
        }

        # Extract customer and project info
        for row_data in e_sheet["data_rows"][:15]:  # First 15 rows contain header info
            for cell in row_data["cells"]:
                value = cell.get("value")
                if value is None:
                    continue

                # Look for customer name (after "수       신:")
                if "수       신:" in str(value):
                    # Next cell should have customer name
                    pass

                # Look for project name (after "건       명:")
                if "건       명:" in str(value):
                    pass

        mapping["cover_sheet"] = cover_mapping

    return mapping


def analyze_panel_grouping(filled_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Analyze how multiple panels and quantities are handled."""
    print("\n[Panel Grouping Analysis]")

    grouping = {
        "total_rows": len(filled_rows),
        "panels": [],
        "patterns": {
            "multiple_panels": False,
            "quantity_over_1": [],
            "section_markers": []
        }
    }

    current_section = None

    for row in filled_rows:
        row_num = row["row"]
        data = row["data"]

        # Get key fields
        품명 = data.get("B", "") or ""
        규격 = data.get("C", "") or ""
        수량 = data.get("E", "") or ""

        print(f"Row {row_num}: {품명} | {규격} | QTY={수량}")

        # Check for section markers (소계, 합계, etc.)
        if 품명 and any(keyword in 품명 for keyword in ["소계", "합계", "TOTAL", "계"]):
            grouping["patterns"]["section_markers"].append({
                "row": row_num,
                "type": 품명
            })
            current_section = None
            continue

        # Check for panel indicators
        if 품명 and any(keyword in 품명 for keyword in ["분전반", "패널", "PANEL", "외함"]):
            current_section = {
                "start_row": row_num,
                "name": 품명,
                "spec": 규격,
                "items": []
            }
            grouping["panels"].append(current_section)

        # Track quantities over 1
        try:
            qty = float(수량) if 수량 else 0
            if qty > 1:
                grouping["patterns"]["quantity_over_1"].append({
                    "row": row_num,
                    "item": 품명,
                    "quantity": qty
                })
        except (ValueError, TypeError):
            pass

        # Add to current section
        if current_section is not None:
            current_section["items"].append({
                "row": row_num,
                "name": 품명,
                "spec": 규격
            })

    grouping["patterns"]["multiple_panels"] = len(grouping["panels"]) > 1

    print(f"\nTotal panels found: {len(grouping['panels'])}")
    print(f"Items with quantity > 1: {len(grouping['patterns']['quantity_over_1'])}")
    print(f"Section markers found: {len(grouping['patterns']['section_markers'])}")

    return grouping


def main():
    template_path = r"C:\Users\PC\Desktop\NABERAL_PROJECT-master\절대코어파일\견적서양식.xlsx"
    example_path = r"C:\Users\PC\Desktop\NABERAL_PROJECT-master\절대코어파일\250922 유광기전.xlsx"

    # Verify files exist
    if not Path(template_path).exists():
        print(f"[ERROR] Template file not found: {template_path}")
        return

    if not Path(example_path).exists():
        print(f"[ERROR] Example file not found: {example_path}")
        return

    # Compare files
    comparison = compare_files(template_path, example_path)

    # Find data mapping
    mapping = find_data_mapping(comparison)

    # Analyze panel grouping
    if mapping["estimate_sheet"].get("filled_rows"):
        grouping = analyze_panel_grouping(mapping["estimate_sheet"]["filled_rows"])
        mapping["panel_grouping"] = grouping

    # Save results
    output_dir = Path("claudedocs")
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "estimate_comparison_analysis.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump({
            "comparison": comparison,
            "mapping": mapping
        }, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] Analysis complete")
    print(f"Output: {output_file}")

    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)

    if "estimate_sheet" in mapping:
        est = mapping["estimate_sheet"]
        print(f"\n[Estimate Sheet Mapping]")
        print(f"  Data range: Row {est.get('data_start_row')} to {est.get('data_end_row')}")
        print(f"  Total items: {est.get('total_items')}")

        if "column_mapping" in est:
            print(f"\n  Column mapping:")
            for col, name in est["column_mapping"].items():
                print(f"    {col}: {name}")

    if "panel_grouping" in mapping:
        pg = mapping["panel_grouping"]
        print(f"\n[Panel Grouping Patterns]")
        print(f"  Multiple panels: {pg['patterns']['multiple_panels']}")
        print(f"  Total panels: {len(pg['panels'])}")
        print(f"  Items with QTY > 1: {len(pg['patterns']['quantity_over_1'])}")

        if pg["patterns"]["quantity_over_1"]:
            print(f"\n  Examples of QTY > 1:")
            for item in pg["patterns"]["quantity_over_1"][:5]:  # Show first 5
                print(f"    Row {item['row']}: {item['item']} x {item['quantity']}")


if __name__ == "__main__":
    main()
