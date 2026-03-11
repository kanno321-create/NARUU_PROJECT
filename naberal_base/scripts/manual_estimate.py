"""
수동 견적 생성 스크립트 - 워크플로우 기반
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

# 견적 데이터
estimate_data = {
    "외함": {
        "품명": "옥외노출 SUS201 1.2T",
        "규격": "600*800*180",
        "단위": "EA",
        "수량": 1,
        "단가": 193000,
    },
    "메인": {
        "품명": "MCCB (SBE-104)",
        "규격": "4P   60AT 14kA",
        "단위": "EA",
        "수량": 1,
        "단가": 20000,
    },
    "분기": [
        {
            "품명": "ELB (SES-54)",
            "규격": "4P   40AT 14kA",
            "단위": "EA",
            "수량": 2,
            "단가": 33400,
        },
        {
            "품명": "ELB (SEC-52)",
            "규격": "2P   40AT  5kA",
            "단위": "EA",
            "수량": 2,
            "단가": 9400,
        },
        {
            "품명": "ELB (SIE-32)",
            "규격": "2P   20AT 2.5kA",
            "단위": "EA",
            "수량": 10,
            "단가": 3800,
        },
    ],
    "부속자재": [
        {"품명": "E.T", "규격": "", "단위": "EA", "수량": 3, "단가": 4500},
        {"품명": "N.T", "규격": "", "단위": "EA", "수량": 1, "단가": 3000},
        {"품명": "N.P", "규격": "CARD HOLDER", "단위": "EA", "수량": 14, "단가": 800},
        {"품명": "N.P", "규격": "3T*40*200", "단위": "EA", "수량": 1, "단가": 4000},
        {"품명": "MAIN BUS-BAR", "규격": "5T*20", "단위": "KG", "수량": 4.2, "단가": 20000},
        {"품명": "BUS-BAR", "규격": "3T*15", "단위": "KG", "수량": 2.5, "단가": 20000},
        {"품명": "COATING", "규격": "PVC(20mm)", "단위": "M", "수량": 1, "단가": 5000},
        {"품명": "P-COVER", "규격": "아크릴(PC)", "단위": "EA", "수량": 1, "단가": 27000},
        {"품명": "ELB지지대", "규격": "", "단위": "EA", "수량": 10, "단가": 500},
        {"품명": "INSULATOR", "규격": "EPOXY 40*40", "단위": "EA", "수량": 4, "단가": 1100},
        {"품명": "잡자재비", "규격": "", "단위": "식", "수량": 1, "단가": 10000},
        {"품명": "ASSEMBLY CHARGE", "규격": "", "단위": "식", "수량": 1, "단가": 95000},
    ]
}

def create_estimate():
    wb = Workbook()

    # 견적서 시트
    ws = wb.active
    ws.title = "견적서"

    # 헤더
    ws["A1"] = " 공종명 : "
    ws["B1"] = "분전반"

    ws["A2"] = "번호"
    ws["B2"] = "품명"
    ws["C2"] = "규격"
    ws["D2"] = "단위"
    ws["E2"] = "수량"
    ws["F2"] = "단   가"
    ws["G2"] = "금   액"

    row = 3
    item_num = 1

    # 1. 외함
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = estimate_data["외함"]["품명"]
    ws[f"C{row}"] = estimate_data["외함"]["규격"]
    ws[f"D{row}"] = estimate_data["외함"]["단위"]
    ws[f"E{row}"] = estimate_data["외함"]["수량"]
    ws[f"F{row}"] = estimate_data["외함"]["단가"]
    ws[f"G{row}"] = estimate_data["외함"]["단가"] * estimate_data["외함"]["수량"]
    row += 1
    item_num += 1

    # 2. 메인차단기
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = estimate_data["메인"]["품명"]
    ws[f"C{row}"] = estimate_data["메인"]["규격"]
    ws[f"D{row}"] = estimate_data["메인"]["단위"]
    ws[f"E{row}"] = estimate_data["메인"]["수량"]
    ws[f"F{row}"] = estimate_data["메인"]["단가"]
    ws[f"G{row}"] = estimate_data["메인"]["단가"] * estimate_data["메인"]["수량"]
    row += 1
    item_num += 1

    # 3. 분기차단기
    for breaker in estimate_data["분기"]:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = breaker["품명"]
        ws[f"C{row}"] = breaker["규격"]
        ws[f"D{row}"] = breaker["단위"]
        ws[f"E{row}"] = breaker["수량"]
        ws[f"F{row}"] = breaker["단가"]
        ws[f"G{row}"] = breaker["단가"] * breaker["수량"]
        row += 1
        item_num += 1

    # 4. 부속자재
    for item in estimate_data["부속자재"]:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = item["품명"]
        ws[f"C{row}"] = item["규격"]
        ws[f"D{row}"] = item["단위"]
        ws[f"E{row}"] = item["수량"]
        ws[f"F{row}"] = item["단가"]
        ws[f"G{row}"] = item["단가"] * item["수량"]
        row += 1
        item_num += 1

    # 빈 행
    for _ in range(30 - row + 1):
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = 0
        row += 1

    # 소계
    ws[f"B{row}"] = "소     계"
    ws[f"D{row}"] = "식"
    ws[f"E{row}"] = 1
    ws[f"G{row}"] = f"=SUM(G3:G{row-1})"
    subtotal_row = row
    row += 1

    # 합계
    ws[f"B{row}"] = "합     계"
    ws[f"D{row}"] = "식"
    ws[f"E{row}"] = 1
    ws[f"G{row}"] = f"=G{subtotal_row}"
    total_row = row

    # 표지 시트
    ws_cover = wb.create_sheet("표지", 0)

    # 표지 헤더
    ws_cover.merge_cells("A1:K1")
    ws_cover["A1"] = "견   적   서"
    ws_cover["A1"].font = Font(size=24, bold=True)
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_cover["A3"] = "일       자: "
    ws_cover["C3"] = datetime.now().strftime("%Y년 %m월 %d일")

    ws_cover["G3"] = "㈜ 한 국 산 업"
    ws_cover["G5"] = "주소:   경기도 안산시 단원구 시화벤처로 467"
    ws_cover["G6"] = "TEL  031-431-1413"
    ws_cover["G7"] = "FAX  031-431-1419"
    ws_cover["G10"] = "대표이사 :  李   忠   源   (인)"

    ws_cover["A5"] = "수       신:           "
    ws_cover["C5"] = "대표님"

    ws_cover["A7"] = "건       명:           "
    ws_cover["C7"] = "워크플로우 기반 견적"

    ws_cover["A11"] = "신한은행 100-031-963940"

    # 합계
    ws_cover["A15"] = " 합   계 : "
    ws_cover["C15"] = f'=CONCATENATE("일금",TEXT(I18,"[DBNum1]"),"원정(V.A.T별도)")'

    # 품목 헤더
    ws_cover["A16"] = "순서"
    ws_cover["B16"] = "품    명"
    ws_cover["D16"] = "규    격(가로*세로*깊이)"
    ws_cover["F16"] = "단위"
    ws_cover["G16"] = "수량"
    ws_cover["H16"] = "단   가"
    ws_cover["I16"] = "금   액"
    ws_cover["J16"] = "비   고"

    # 분전반 품목
    ws_cover["A17"] = 1
    ws_cover["B17"] = "분전반"
    ws_cover["D17"] = "600*800*180"
    ws_cover["F17"] = "면"
    ws_cover["G17"] = 1
    ws_cover["H17"] = f"=+견적서!G{total_row}"
    ws_cover["I17"] = "=H17*G17"
    ws_cover["J17"] = "옥외노출"

    # 합계 행
    ws_cover["B18"] = "     합         계"
    ws_cover["F18"] = "식"
    ws_cover["G18"] = "=SUM(G17:G17)"
    ws_cover["I18"] = "=SUM(I17:I17)"

    # 부가세
    ws_cover["H19"] = "부가세포함"
    ws_cover["I19"] = "=I18*1.1"

    # NOTE
    ws_cover["A20"] = "< NOTE > 1. 차단기: 상도차단기    , 외함: 옥외노출 SUS201 1.2T (HDS)"
    ws_cover["A21"] = "             2. 제작도면 및 승인도면 제작비 별도"
    ws_cover["A22"] = "             3. 운임비 별도 (화물배송, 택배배송, 1톤 용달)"
    ws_cover["A24"] = "결제조건:200만원이하 출고전 완불/ 200만원이상 계약금30% 출고전 완불"

    # 파일 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"C:\\Users\\PC\\Desktop\\NABERAL_PROJECT-master\\outputs\\estimate_workflow_{timestamp}.xlsx"
    wb.save(filename)
    print(f"[OK] 견적서 생성 완료: {filename}")
    return filename

if __name__ == "__main__":
    create_estimate()
