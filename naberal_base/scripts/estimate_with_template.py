"""
견적서 양식 기반 견적 생성 스크립트
"""
import shutil
from datetime import datetime
from openpyxl import load_workbook

# 견적 데이터
estimate_data = {
    "외함": {
        "품명": "옥외노출 SUS201 1.2T",
        "규격": "600*800*180",
        "단위": "EA",
        "수량": 1,
        "단가": 193000,
    },
    "메인": {
        "품명": "MCCB (SBE-104)",
        "규격": "4P   60AT 14kA",
        "단위": "EA",
        "수량": 1,
        "단가": 20000,
    },
    "분기": [
        {
            "품명": "ELB (SES-54)",
            "규격": "4P   40AT 14kA",
            "단위": "EA",
            "수량": 2,
            "단가": 33400,
        },
        {
            "품명": "ELB (SEC-52)",
            "규격": "2P   40AT  5kA",
            "단위": "EA",
            "수량": 2,
            "단가": 9400,
        },
        {
            "품명": "ELB (SIE-32)",
            "규격": "2P   20AT 2.5kA",
            "단위": "EA",
            "수량": 10,
            "단가": 3800,
        },
    ],
    "부속자재": [
        {"품명": "E.T", "규격": "", "단위": "EA", "수량": 2, "단가": 4500},
        {"품명": "N.T", "규격": "", "단위": "EA", "수량": 1, "단가": 3000},
        {"품명": "N.P", "규격": "CARD HOLDER", "단위": "EA", "수량": 14, "단가": 800},
        {"품명": "N.P", "규격": "3T*40*200", "단위": "EA", "수량": 1, "단가": 4000},
        {"품명": "MAIN BUS-BAR", "규격": "3T*15", "단위": "KG", "수량": 2.9, "단가": 20000},
        {"품명": "BUS-BAR", "규격": "3T*15", "단위": "KG", "수량": 2.5, "단가": 20000},
        {"품명": "COATING", "규격": "PVC(20mm)", "단위": "M", "수량": 1, "단가": 5000},
        {"품명": "P-COVER", "규격": "아크릴(PC)", "단위": "EA", "수량": 1, "단가": 19200},
        {"품명": "ELB지지대", "규격": "", "단위": "EA", "수량": 10, "단가": 500},
        {"품명": "INSULATOR", "규격": "EPOXY 40*40", "단위": "EA", "수량": 4, "단가": 1100},
        {"품명": "잡자재비", "규격": "", "단위": "식", "수량": 1, "단가": 10000},
        {"품명": "ASSEMBLY CHARGE", "규격": "", "단위": "식", "수량": 1, "단가": 75000},
    ]
}

def create_estimate():
    # 원본 양식 복사
    template_path = r"C:\Users\PC\Desktop\NABERAL_PROJECT-master\절대코어파일\견적서양식.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"C:\\Users\\PC\\Desktop\\NABERAL_PROJECT-master\\outputs\\estimate_{timestamp}.xlsx"

    shutil.copy(template_path, output_path)

    # 파일 열기
    wb = load_workbook(output_path)
    ws = wb["견적서"]
    ws_cover = wb["표지"]

    row = 3
    item_num = 1

    # 1. 외함
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = estimate_data["외함"]["품명"]
    ws[f"C{row}"] = estimate_data["외함"]["규격"]
    ws[f"D{row}"] = estimate_data["외함"]["단위"]
    ws[f"E{row}"] = estimate_data["외함"]["수량"]
    ws[f"F{row}"] = estimate_data["외함"]["단가"]
    ws[f"G{row}"] = estimate_data["외함"]["단가"] * estimate_data["외함"]["수량"]
    row += 1
    item_num += 1

    # 2. 메인차단기
    ws[f"A{row}"] = item_num
    ws[f"B{row}"] = estimate_data["메인"]["품명"]
    ws[f"C{row}"] = estimate_data["메인"]["규격"]
    ws[f"D{row}"] = estimate_data["메인"]["단위"]
    ws[f"E{row}"] = estimate_data["메인"]["수량"]
    ws[f"F{row}"] = estimate_data["메인"]["단가"]
    ws[f"G{row}"] = estimate_data["메인"]["단가"] * estimate_data["메인"]["수량"]
    row += 1
    item_num += 1

    # 3. 분기차단기
    for breaker in estimate_data["분기"]:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = breaker["품명"]
        ws[f"C{row}"] = breaker["규격"]
        ws[f"D{row}"] = breaker["단위"]
        ws[f"E{row}"] = breaker["수량"]
        ws[f"F{row}"] = breaker["단가"]
        ws[f"G{row}"] = breaker["단가"] * breaker["수량"]
        row += 1
        item_num += 1

    # 4. 부속자재
    for item in estimate_data["부속자재"]:
        ws[f"A{row}"] = item_num
        ws[f"B{row}"] = item["품명"]
        ws[f"C{row}"] = item["규격"]
        ws[f"D{row}"] = item["단위"]
        ws[f"E{row}"] = item["수량"]
        ws[f"F{row}"] = item["단가"]
        ws[f"G{row}"] = item["단가"] * item["수량"]
        row += 1
        item_num += 1

    # 소계/합계 수식 업데이트
    ws["G48"] = f"=SUM(G3:G{row-1})"
    ws["G49"] = f"=G48"

    # 표지 업데이트
    ws_cover["C3"] = datetime.now().strftime("%Y년 %m월 %d일")

    # 18~32번 빈 행 삭제 (분전반이 1개뿐이므로)
    ws_cover.delete_rows(18, 15)

    # 17번 행: 분전반 데이터
    ws_cover["A17"] = 1
    ws_cover["B17"] = "분전반"
    ws_cover["D17"] = "600*800*180"
    ws_cover["F17"] = "면"
    ws_cover["G17"] = 1
    ws_cover["H17"] = "=+견적서!G49"
    ws_cover["I17"] = "=H17*G17"
    ws_cover["J17"] = "옥외노출"

    # 18번 행: 합계 (삭제 후 33번이 18번으로 이동)
    ws_cover["A18"] = None
    ws_cover["B18"] = "     합         계"
    ws_cover["F18"] = "식"
    ws_cover["G18"] = "=SUM(G17:G17)"
    ws_cover["I18"] = "=SUM(I17:I17)"

    # 19번 행: 부가세포함 (삭제 후 34번이 19번으로 이동)
    ws_cover["A19"] = None
    ws_cover["H19"] = "부가세포함"
    ws_cover["I19"] = "=I18*1.1"

    # 20번 행: NOTE (삭제 후 35번이 20번으로 이동)
    ws_cover["A20"] = '< NOTE > 1. 차단기:   상도차단기          , 외함:  옥외SUS201 1.2T (SC)'

    # 표지 합계 텍스트 업데이트 (18번 행 참조)
    ws_cover["C15"] = '=CONCATENATE("일금",TEXT(I18,"[DBNum1]"),"원정(V.A.T별도)")'

    # 저장
    wb.save(output_path)
    print(f"[OK] 견적서 생성 완료: {output_path}")
    return output_path

if __name__ == "__main__":
    create_estimate()
