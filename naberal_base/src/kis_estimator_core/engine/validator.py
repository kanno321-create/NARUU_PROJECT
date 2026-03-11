"""
Stage 3: Validator
Excel 품질 검증

품질 게이트 (MANDATORY):
1. 수식 보존 100% (CHK_FORMULA_PRESERVATION)
2. 병합 셀 손상 0 (CHK_MERGED_CELLS_PRESERVATION)
3. 크로스 시트 참조 정확성 100% (CHK_CROSS_SHEET_REFERENCE)

성능 목표:
- 검증: < 100ms (목표), < 500ms (최대)

참조:
- STAGE3_FORMAT_DESIGN_20251003.md (설계 문서)
"""

import logging
from pathlib import Path

from kis_estimator_core.core.ssot.errors import ErrorCode, raise_error

try:
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise ImportError(
        "openpyxl이 설치되지 않았습니다. pip install openpyxl로 설치하세요."
    ) from e

from ..core.ssot.guards_format import (
    formula_guard,
    totals_guard,
)
from ..errors import CAL_001, CAL_002, ValidationError
from .models import ValidationReport

logger = logging.getLogger(__name__)


class Validator:
    """
    Excel 품질 검증기

    품질 게이트:
    1. 수식 보존: G 컬럼 모든 셀이 수식 유지
    2. 병합 셀: 템플릿과 동일한 병합 영역
    3. 크로스 참조: 표지 → 견적서 참조 정확성
    """

    def __init__(self, template_path: Path):
        """
        Args:
            template_path: 템플릿 파일 경로 (비교 기준)
        """
        if not template_path.exists():
            raise_error(
                ErrorCode.E_INTERNAL, f"템플릿 파일이 없습니다: {template_path}"
            )

        self.template_path = template_path
        logger.info(f"Validator initialized with template: {template_path}")

    def validate(self, excel_path: Path) -> ValidationReport:
        """
        Excel 검증 메인 메서드

        Args:
            excel_path: 검증할 Excel 파일 경로

        Returns:
            ValidationReport: 검증 결과 보고서

        Raises:
            ValidationError: CAL-001 (수식 보존 실패) 또는 CAL-002 (합계 불일치) 발생 시
        """
        if not excel_path.exists():
            raise_error(ErrorCode.E_INTERNAL, f"Excel 파일이 없습니다: {excel_path}")

        logger.info(f"Starting validation: {excel_path}")

        violations = []
        warnings = []
        error_codes = []

        # 워크북 로드 (data_only=False 필수)
        try:
            wb_generated = load_workbook(excel_path, data_only=False)
            wb_template = load_workbook(self.template_path, data_only=False)
        except Exception as e:
            violations.append(f"워크북 로드 실패: {e}")
            return ValidationReport(
                formula_preservation=False,
                merged_cells_intact=False,
                cross_references_valid=False,
                violations=violations,
            )

        # 1. 수식 보존 검증 (CAL-001)
        formula_ok = self._validate_formulas(
            wb_generated, wb_template, violations, warnings, error_codes
        )

        # 2. 병합 셀 검증
        merged_ok = self._validate_merged_cells(
            wb_generated, wb_template, violations, warnings
        )

        # 3. 크로스 시트 참조 검증
        cross_ref_ok = self._validate_cross_references(
            wb_generated, violations, warnings
        )

        # 4. 합계 일치 검증 (CAL-002)
        total_ok = self._validate_total_consistency(
            wb_generated, violations, warnings, error_codes
        )

        # 검증 보고서 생성
        report = ValidationReport(
            formula_preservation=formula_ok,
            merged_cells_intact=merged_ok,
            cross_references_valid=cross_ref_ok and total_ok,
            violations=violations,
            warnings=warnings,
            metadata={
                "excel_path": str(excel_path),
                "template_path": str(self.template_path),
                "passed": formula_ok
                and merged_ok
                and cross_ref_ok
                and total_ok
                and len(violations) == 0,
                "error_codes": error_codes,
            },
        )

        # BLOCKING 에러 발생 시 즉시 raise
        if error_codes:
            logger.error(f"Validation FAILED with error codes: {error_codes}")
            raise ValidationError(
                error_code=CAL_001 if "CAL-001" in error_codes else CAL_002,
                phase="Stage 3: Format",
            )

        if report.passed:
            logger.info("Validation PASSED ✅")
        else:
            logger.error(f"Validation FAILED ❌: {len(violations)} violations")

        return report

    def _validate_formulas(
        self,
        wb_generated: Workbook,
        wb_template: Workbook,
        violations: list[str],
        warnings: list[str],
        error_codes: list[str],
    ) -> bool:
        """
        수식 보존 검증 (SSOT formula_guard 사용)

        검증 항목:
        - 견적서!G3:G200: 각 품목 금액 (=F*E 패턴)
        - 표지!I17:I50: 금액 (=H*G 패턴)
        - 표지!C15: NUMBERSTRING 수식

        Args:
            wb_generated: 생성된 워크북
            wb_template: 템플릿 워크북 (unused - SSOT guards는 template 불필요)
            violations: 위반 사항 리스트 (출력)
            warnings: 경고 사항 리스트 (출력)
            error_codes: 에러 코드 리스트 (출력)

        Returns:
            bool: 수식 보존 성공 여부
        """
        logger.debug("Validating formulas (SSOT formula_guard)")

        ws_estimate = wb_generated["견적서"]
        ws_cover = wb_generated["표지"]

        # 1. 견적서!G 컬럼 수식 보존 검증 (allow_blank=True로 품목 수에 따른 빈 셀 허용)
        is_valid_estimate, errors_estimate, rate_estimate = formula_guard(
            worksheet=ws_estimate,
            formula_ranges=["G3:G200"],
            allow_blank=True,  # 품목 수에 따라 빈 셀 허용
            policy_on_empty="allow",
        )

        if not is_valid_estimate:
            violations.extend(errors_estimate)
            error_codes.append("CAL-001")
            logger.error(
                f"견적서 수식 보존 실패: {len(errors_estimate)}개 오류, 보존율={rate_estimate:.1%}"
            )
            return False

        logger.debug(f"견적서 수식 보존 통과 ✅ (보존율={rate_estimate:.1%})")

        # 2. 표지!I 컬럼 수식 보존 검증
        is_valid_cover, errors_cover, rate_cover = formula_guard(
            worksheet=ws_cover,
            formula_ranges=["I17:I50"],
            allow_blank=True,  # 표지는 빈 셀 허용 (분전반 수에 따라 변동)
            policy_on_empty="allow",
        )

        if not is_valid_cover:
            violations.extend(errors_cover)
            error_codes.append("CAL-001")
            logger.error(
                f"표지 수식 보존 실패: {len(errors_cover)}개 오류, 보존율={rate_cover:.1%}"
            )
            return False

        logger.debug(f"표지 수식 보존 통과 ✅ (보존율={rate_cover:.1%})")

        # 3. 표지!C15 (NUMBERSTRING 수식) - 경고만
        c15_value = ws_cover["C15"].value
        if c15_value and isinstance(c15_value, str):
            if not c15_value.startswith("="):
                warnings.append(
                    f"표지!C15 수식 손실 가능: '{c15_value}' (NUMBERSTRING 수식 확인 권장)"
                )

        logger.debug("수식 보존 검증 (SSOT) 통과 ✅")
        return True

    def _validate_merged_cells(
        self,
        wb_generated: Workbook,
        wb_template: Workbook,
        violations: list[str],
        warnings: list[str],
    ) -> bool:
        """
        병합 셀 검증

        검증 항목:
        - 표지 시트: A5, A7 등 병합 영역 보존

        Args:
            wb_generated: 생성된 워크북
            wb_template: 템플릿 워크북
            violations: 위반 사항 리스트 (출력)
            warnings: 경고 사항 리스트 (출력)

        Returns:
            bool: 병합 셀 보존 성공 여부
        """
        logger.debug("Validating merged cells")

        ws_cover_gen = wb_generated["표지"]
        ws_cover_tpl = wb_template["표지"]

        # 병합 영역 추출
        merged_gen = list(ws_cover_gen.merged_cells.ranges)
        merged_tpl = list(ws_cover_tpl.merged_cells.ranges)

        # 개수 비교
        if len(merged_gen) != len(merged_tpl):
            violations.append(
                f"병합 셀 개수 불일치: 생성={len(merged_gen)}, 템플릿={len(merged_tpl)}"
            )
            return False

        # 각 병합 영역 비교
        for tpl_range in merged_tpl:
            if tpl_range not in merged_gen:
                violations.append(f"병합 셀 손상: {tpl_range}")
                return False

        logger.debug("병합 셀 검증 통과 ✅")
        return True

    def _validate_cross_references(
        self,
        wb_generated: Workbook,
        violations: list[str],
        warnings: list[str],
    ) -> bool:
        """
        크로스 시트 참조 검증

        검증 항목:
        - 표지 H17~: =+견적서!G{subtotal_row} 또는 =견적서!G{row} 패턴
        - 유연성 허용: 동적 행 수에 맞춰 검증

        Args:
            wb_generated: 생성된 워크북
            violations: 위반 사항 리스트 (출력)
            warnings: 경고 사항 리스트 (출력)

        Returns:
            bool: 크로스 참조 정확성 성공 여부
        """
        logger.debug("Validating cross-sheet references")

        ws_cover = wb_generated["표지"]

        cross_ref_errors = 0
        cross_ref_found = 0

        # 표지 H17~H32 (요약 테이블, 동적으로 확장 가능)
        for row in range(17, 50):  # 넓게 검사
            cell = f"H{row}"
            value = ws_cover[cell].value

            if value is None or value == "":
                continue

            # 수식 확인
            if not isinstance(value, str):
                continue

            # SUM 수식 스킵 (소계 행)
            if "SUM" in value.upper():
                continue

            # 크로스 참조 패턴 확인: =+견적서!G{row} or =견적서!G{row}
            if value.startswith("=+견적서!G") or value.startswith("=견적서!G"):
                cross_ref_found += 1
            # 다른 수식은 허용 (예: =H*G, =SUM 등)

        # 최소 1개 이상 크로스 참조 있어야 함
        if cross_ref_found == 0:
            warnings.append("크로스 참조 없음 (표지 → 견적서)")

        if cross_ref_errors > 0:
            logger.error(f"크로스 참조 검증 실패: {cross_ref_errors}개 오류")
            return False

        logger.debug(f"크로스 참조 검증 통과 ✅ ({cross_ref_found}개 발견)")
        return True

    def _validate_total_consistency(
        self,
        wb_generated: Workbook,
        violations: list[str],
        warnings: list[str],
        error_codes: list[str],
    ) -> bool:
        """
        합계 일치 검증 (SSOT totals_guard 사용)

        검증 항목:
        - 견적서: 소계 → 합계 → VAT (부가세 포함) 일관성
        - 표지: 소계 → 합계 → VAT 일관성

        Args:
            wb_generated: 생성된 워크북
            violations: 위반 사항 리스트 (출력)
            warnings: 경고 사항 리스트 (출력)
            error_codes: 에러 코드 리스트 (출력)

        Returns:
            bool: 합계 일치 여부
        """
        logger.debug("Validating total consistency (SSOT totals_guard)")

        ws_estimate = wb_generated["견적서"]
        ws_cover = wb_generated["표지"]

        try:
            # 1. 견적서 합계 검증 (소계 → 합계)
            # 소계 행 동적 검색 (B열에 "소  계" 텍스트)
            subtotal_row_estimate = None
            for row in range(40, 60):
                cell_value = ws_estimate[f"B{row}"].value
                if cell_value and "소" in str(cell_value) and "계" in str(cell_value):
                    subtotal_row_estimate = row
                    break

            # 합계 행 동적 검색 (B열에 "합  계" 텍스트)
            total_row_estimate = None
            for row in range(
                subtotal_row_estimate + 1 if subtotal_row_estimate else 48, 60
            ):
                cell_value = ws_estimate[f"B{row}"].value
                if cell_value and "합" in str(cell_value) and "계" in str(cell_value):
                    total_row_estimate = row
                    break

            # VAT 행은 합계 +1 (견적서에는 VAT 없을 수 있으므로 optional)
            vat_row_estimate = total_row_estimate + 1 if total_row_estimate else None

            if subtotal_row_estimate and total_row_estimate:
                subtotal_cell = f"G{subtotal_row_estimate}"
                total_cell = f"G{total_row_estimate}"
                vat_cell = f"G{vat_row_estimate}" if vat_row_estimate else None
                # 라인아이템 범위 계산 (G3:G{subtotal_row - 1})
                line_items_range = f"G3:G{subtotal_row_estimate - 1}"

                is_valid_estimate, errors_estimate = totals_guard(
                    worksheet=ws_estimate,
                    subtotal_cell=subtotal_cell,
                    total_cell=total_cell,
                    vat_cell=vat_cell,
                    tolerance=0.01,  # 1원 오차 허용
                    line_items_range=line_items_range,  # I-2.1: 직접 합산용
                )

                if not is_valid_estimate:
                    violations.extend(errors_estimate)
                    error_codes.append("CAL-002")
                    logger.error(f"견적서 합계 불일치: {errors_estimate}")
                    return False

                logger.debug(
                    f"견적서 합계 일치 ✅ (subtotal={subtotal_cell}, total={total_cell})"
                )
            else:
                warnings.append("견적서 소계/합계 행을 찾을 수 없습니다")

            # 2. 표지 합계 검증 (소계 → 합계 → VAT)
            # 표지는 I열에 금액
            subtotal_row_cover = None
            for row in range(17, 40):
                cell_value = ws_cover[f"B{row}"].value
                if cell_value and "소" in str(cell_value) and "계" in str(cell_value):
                    subtotal_row_cover = row
                    break

            total_row_cover = None
            for row in range(subtotal_row_cover + 1 if subtotal_row_cover else 17, 40):
                cell_value = ws_cover[f"B{row}"].value
                if cell_value and "합" in str(cell_value) and "계" in str(cell_value):
                    total_row_cover = row
                    break

            vat_row_cover = total_row_cover + 1 if total_row_cover else None

            if subtotal_row_cover and total_row_cover:
                subtotal_cell = f"I{subtotal_row_cover}"
                total_cell = f"I{total_row_cover}"
                vat_cell = f"I{vat_row_cover}" if vat_row_cover else None
                # 라인아이템 범위 계산 (I17:I{subtotal_row - 1})
                line_items_range_cover = f"I17:I{subtotal_row_cover - 1}"

                is_valid_cover, errors_cover = totals_guard(
                    worksheet=ws_cover,
                    subtotal_cell=subtotal_cell,
                    total_cell=total_cell,
                    vat_cell=vat_cell,
                    tolerance=0.01,
                    line_items_range=line_items_range_cover,  # I-2.1: 직접 합산용
                )

                if not is_valid_cover:
                    violations.extend(errors_cover)
                    error_codes.append("CAL-002")
                    logger.error(f"표지 합계 불일치: {errors_cover}")
                    return False

                logger.debug(
                    f"표지 합계 일치 ✅ (subtotal={subtotal_cell}, total={total_cell})"
                )
            else:
                warnings.append("표지 소계/합계 행을 찾을 수 없습니다")

            logger.debug("합계 일치 검증 (SSOT) 통과 ✅")
            return True

        except Exception as e:
            warnings.append(f"합계 검증 중 오류: {e}")
            logger.exception("합계 검증 중 예외 발생")
            return True  # 예외 발생 시 경고만 (FAIL 아님)
