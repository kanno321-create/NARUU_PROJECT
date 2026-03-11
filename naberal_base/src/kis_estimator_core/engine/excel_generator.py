"""
Stage 3: Excel Generator (Real Template Mode)
EstimateData → Excel 파일 생성

SPEC KIT 기준 (MANDATORY):
- 실물 템플릿 로드 (견적서양식.xlsx)
- 기존 구조 100% 보존 (회사명, 인감, 주소, NOTE)
- 데이터만 채워넣기 (수신, 건명, 품목 리스트)
- 수식 보존 100% (G 컬럼 입력 금지)
- 크로스 시트 참조 정확성 100%

성능 목표:
- Excel 생성: < 1s (목표), < 3s (최대)

근거:
- Excel MCP 서버 사용 (@negokaz/excel-mcp-server)
- 실물 템플릿 분석 완료 (표지 A1:K5274, 견적서 A1:G49)
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path

from kis_estimator_core.core.ssot.constants_format import (
    SHEET_CONDITIONS,
    UNIT_ASSEMBLY,
    UNIT_ENCLOSURE,
)
from kis_estimator_core.core.ssot.errors import ErrorCode, raise_error
from kis_estimator_core.core.ssot.ssot_loader import load_rounding

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    raise ImportError(
        "openpyxl이 설치되지 않았습니다. pip install openpyxl로 설치하세요."
    ) from e

from .models import EstimateData, PanelEstimate

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Excel 견적서 생성기 (실물 템플릿 로드)

    핵심 원칙:
    1. 실물 템플릿 로드 (견적서양식.xlsx)
    2. 기존 구조 100% 보존 (회사명, 인감, 주소, NOTE)
    3. 데이터만 채워넣기 (수신, 건명, 품목 리스트)
    4. 수식 보존 100% (G 컬럼 입력 금지)
    5. 크로스 시트 참조 동적 생성
    """

    def __init__(self, template_path: Path):
        """
        Args:
            template_path: 실물 템플릿 경로 (견적서양식.xlsx)

        Raises:
            FileNotFoundError: 템플릿 파일 없음 (목업 절대 금지)
        """
        if not template_path.exists():
            raise_error(
                ErrorCode.E_IO,
                f"실물 템플릿 파일을 찾을 수 없습니다: {template_path}\n"
                "목업 절대 금지! 실제 템플릿 파일을 제공하세요.",
            )

        self.template_path = template_path
        logger.info(f"ExcelGenerator initialized: {template_path}")

    def generate(self, data: EstimateData, output_path: Path) -> Path:
        """
        Excel 견적서 생성 메인 메서드

        Args:
            data: EstimateData (변환된 견적 데이터)
            output_path: 출력 파일 경로

        Returns:
            Path: 생성된 Excel 파일 경로

        Raises:
            ValueError: 데이터 검증 실패
        """
        logger.info(f"Starting Excel generation (template mode): {output_path}")

        # 1. 템플릿 복사
        shutil.copy(self.template_path, output_path)
        logger.info(f"Template copied: {self.template_path} → {output_path}")

        # 2. 템플릿 로드 (data_only=False, 수식 보존)
        wb = load_workbook(output_path, data_only=False)
        ws_cover = wb["표지"]
        ws_estimate = wb["견적서"]

        # 3. 견적서 시트 작성 (데이터만)
        subtotal_rows = self._write_estimate_sheet(data.panels, ws_estimate)

        # 4. 표지 시트 작성 (데이터만)
        self._write_cover_sheet(data, ws_cover, subtotal_rows)

        # 5. 조건 시트 작성 (Phase VII: 자동삽입 고도화)
        if SHEET_CONDITIONS in wb.sheetnames:
            # 템플릿에 조건 시트가 있으면 데이터 작성
            ws_conditions = wb[SHEET_CONDITIONS]
            self._write_conditions_sheet(ws_conditions)
            logger.info("Conditions sheet updated (template existing)")
        else:
            # 템플릿에 조건 시트가 없으면 새로 생성
            ws_conditions = wb.create_sheet(SHEET_CONDITIONS)
            self._write_conditions_sheet(ws_conditions)
            logger.info("Conditions sheet created and written")

        # 6. 저장
        wb.save(output_path)
        logger.info(f"Excel generation completed: {output_path}")

        return output_path

    def _write_estimate_sheet(
        self, panels: list[PanelEstimate], ws: Worksheet
    ) -> list[int]:
        """
        견적서 시트 작성 (실물 템플릿에 데이터만 채우기)

        템플릿 구조:
        - Row 1: 공종명 (분전반) - 보존
        - Row 2: 헤더 - 보존
        - Row 3~47: 데이터 행 (G 컬럼: =F*E 수식) - 채우기
        - Row 48: 소계 (=SUM(G3:G47)) - 보존
        - Row 49: 합계 (=G48*E49) - 보존

        Args:
            panels: 분전반 리스트 (현재는 1개만 지원)
            ws: 견적서 시트

        Returns:
            List[int]: 각 분전반 소계 행 번호 (표지 참조용)
        """
        logger.debug("Writing estimate sheet (data only)")

        # 현재는 1개 분전반만 지원
        if len(panels) != 1:
            raise_error(
                ErrorCode.E_NOT_IMPLEMENTED,
                f"현재는 1개 분전반만 지원합니다. 입력: {len(panels)}개",
                hint="다중 분전반 지원은 Phase 4에서 구현 예정",
            )

        panel = panels[0]
        all_items = panel.all_items_sorted

        # Row 3부터 데이터 채우기
        current_row = 3

        for _item_idx, item in enumerate(all_items):
            # A: 번호 (자동 생성하지 않음, 템플릿 기본값 사용)

            # B: 품명
            ws[f"B{current_row}"] = item.item_name

            # C: 규격
            ws[f"C{current_row}"] = item.spec

            # D: 단위
            ws[f"D{current_row}"] = item.unit

            # E: 수량
            ws[f"E{current_row}"] = item.quantity

            # F: 단가
            ws[f"F{current_row}"] = item.unit_price

            # G: 금액 (수식 보존 - 입력하지 않음!)

            current_row += 1

            # 템플릿 행 개수 초과 방지
            if current_row > 47:
                logger.warning(
                    f"품목 개수 초과: {len(all_items)}개 > 45개 (템플릿 한계)"
                )
                break

        # 소계 행 번호 반환 (Row 48)
        subtotal_rows = [48]

        logger.info(
            f"Estimate sheet written: {len(all_items)} items, subtotal at Row 48"
        )
        return subtotal_rows

    def _write_cover_sheet(
        self,
        data: EstimateData,
        ws_cover: Worksheet,
        subtotal_rows: list[int],
    ):
        """
        표지 시트 작성 (견적서표지작성법.txt 준수)

        표지작성법 규칙:
        1. C3: 날짜 (YYYY년 MM월 DD일)
        2. C5: 거래처명
        3. C7: 건명 (없으면 공란)
        4. B17~: 분전반 정보
           - B: 분전반명 (없으면 "분전반", 여러개면 "분전반1", "분전반2")
           - D: 외함 사이즈
           - F: 단위 (UNIT_ENCLOSURE)
           - G: 수량
           - H: 견적가 (=+견적서!G[소계행])
           - J: 특이사항 (속판제작, 옥내노출 등)
        5. 합계 처리:
           - 1종류: B18="합   계", I18=SUM(I17:I17), I19=I18*1.1
           - 2종류+: 마지막 하단에 "합   계", SUM, 부가세

        Args:
            data: EstimateData
            ws_cover: 표지 시트
            subtotal_rows: 각 분전반 소계 행 번호
        """
        logger.debug("Writing cover sheet following 견적서표지작성법.txt")

        # 1. 날짜 (C3) - YYYY년 MM월 DD일 형식
        today = datetime.now()
        ws_cover["C3"] = f"{today.year}년 {today.month:02d}월 {today.day:02d}일"

        # 2. 거래처명 (C5)
        ws_cover["C5"] = data.customer_name

        # 3. 건명 (C7) - 없으면 공란
        ws_cover["C7"] = data.project_name if data.project_name else ""

        # 4. 분전반 정보 (Row 17부터)
        if len(data.panels) != 1:
            raise_error(
                ErrorCode.E_NOT_IMPLEMENTED,
                f"현재는 1개 분전반만 지원합니다. 입력: {len(data.panels)}개",
                hint="다중 분전반 지원은 Phase 4에서 구현 예정",
            )

        panel = data.panels[0]
        row = 17

        # B17: 분전반명 (없으면 "분전반")
        panel_name = getattr(panel, "name", None) or "분전반"
        ws_cover[f"B{row}"] = panel_name

        # D17: 외함 사이즈
        if panel.enclosure:
            ws_cover[f"D{row}"] = panel.enclosure.spec
        else:
            ws_cover[f"D{row}"] = "N/A"

        # F17: 단위 (UNIT_ENCLOSURE)
        ws_cover[f"F{row}"] = UNIT_ENCLOSURE

        # G17: 수량 (일반적으로 1)
        ws_cover[f"G{row}"] = 1

        # H17: 견적가 (크로스 시트 참조)
        subtotal_row = subtotal_rows[0]
        ws_cover[f"H{row}"] = f"=+견적서!G{subtotal_row}"

        # I17: 금액 (=H17*G17)
        ws_cover[f"I{row}"] = f"=H{row}*G{row}"

        # J17: 특이사항 (외함 종류)
        enclosure_type = (
            getattr(panel.enclosure, "item_name", None) if panel.enclosure else "N/A"
        )
        ws_cover[f"J{row}"] = enclosure_type

        # 5. 합계 처리 (1종류)
        # B18: "합   계"
        ws_cover["B18"] = "합   계"
        # F18: 단위 UNIT_ASSEMBLY
        ws_cover["F18"] = UNIT_ASSEMBLY
        # G18: 수량 (=G17)
        ws_cover["G18"] = f"=G{row}"
        # I18: 합계 (=SUM(I17:I17))
        ws_cover["I18"] = f"=SUM(I{row}:I{row})"
        # I19: 부가세 포함 (SSOT rounding.json vat_pct 사용)
        rounding = load_rounding()
        vat_multiplier = 1 + rounding["vat_pct"]  # 1.10 from SSOT
        ws_cover["I19"] = f"=I18*{vat_multiplier}"

        logger.info(
            f"Cover sheet written: {panel_name}, size={panel.enclosure.spec if panel.enclosure else 'N/A'}"
        )

    def _write_conditions_sheet(self, ws_conditions: Worksheet):
        """
        조건 시트 작성 (견적조건표)

        조건 시트 구조:
        - B2: 제목 "견적조건"
        - B4~B13: 조건 번호 (1~10)
        - C4~C13: 조건 설명 (CONDITION_1~10)

        Args:
            ws_conditions: 조건 시트

        Note:
            SSOT 상수 (constants_format.py) 사용:
            - CONDITIONS_TITLE_CELL, CONDITIONS_START_ROW
            - CONDITION_1_PANEL_NAME ~ CONDITION_10_MULTI_PANELS
        """
        from kis_estimator_core.core.ssot.constants_format import (
            CONDITION_1_PANEL_NAME,
            CONDITION_2_ENCLOSURE_QTY,
            CONDITION_3_MAIN_QTY,
            CONDITION_4_BREAKER_SPEC,
            CONDITION_5_PANEL_SPEC,
            CONDITION_6_BRANCH_TYPES,
            CONDITION_7_REQUIRED_MATERIALS,
            CONDITION_8_BLANK_ROWS,
            CONDITION_9_매입함,
            CONDITION_10_MULTI_PANELS,
            CONDITIONS_DESC_COL,
            CONDITIONS_ITEM_COL,
            CONDITIONS_START_ROW,
            CONDITIONS_TITLE_CELL,
        )

        logger.debug("Writing conditions sheet")

        # 제목
        ws_conditions[CONDITIONS_TITLE_CELL] = "견적조건"

        # 10가지 조건 항목
        conditions = [
            CONDITION_1_PANEL_NAME,
            CONDITION_2_ENCLOSURE_QTY,
            CONDITION_3_MAIN_QTY,
            CONDITION_4_BREAKER_SPEC,
            CONDITION_5_PANEL_SPEC,
            CONDITION_6_BRANCH_TYPES,
            CONDITION_7_REQUIRED_MATERIALS,
            CONDITION_8_BLANK_ROWS,
            CONDITION_9_매입함,
            CONDITION_10_MULTI_PANELS,
        ]

        for i, condition_text in enumerate(conditions, start=1):
            row = CONDITIONS_START_ROW + (i - 1)
            # 조건 번호
            ws_conditions[f"{CONDITIONS_ITEM_COL}{row}"] = f"{i}."
            # 조건 설명
            ws_conditions[f"{CONDITIONS_DESC_COL}{row}"] = condition_text

        logger.info(f"Conditions sheet written: {len(conditions)} items")
