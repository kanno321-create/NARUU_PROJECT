"""
Stage 5: Doc Lint Guard
최종 문서 품질 검증

품질 게이트 (MANDATORY):
1. 수식 보존 = 100%
2. 크로스 참조 = 100%
3. 필수 자재 12개 누락 = 0
4. 린트 오류 = 0

성능 목표:
- 검증: < 200ms (목표), < 1s (최대)

참조:
- CLAUDE.md (견적조건 7번)
- validator.py (Stage 3 검증)
"""

import logging
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    raise ImportError(
        "openpyxl이 설치되지 않았습니다. pip install openpyxl로 설치하세요."
    ) from e

from .models import ValidationReport

logger = logging.getLogger(__name__)


class DocLintGuard:
    """
    최종 문서 품질 검증기 (Stage 5)

    역할:
    - Stage 3 validator 결과 종합
    - 견적조건 7번 필수 자재 12개 누락 체크
    - 최종 품질 점수 계산
    - 린트 오류 보고
    """

    # 견적조건 7번: 필수 자재 12개 (순서대로)
    REQUIRED_ITEMS = [
        "E.T",
        "N.T",
        "N.P",  # CARD HOLDER
        "MAIN BUS-BAR",
        "BUS-BAR",  # 분기용 (일반 견적) 또는 부스바 처리 옵션 (메인만)
        "COATING",
        "P-COVER",
        "차단기지지대",  # 400AF 이상만
        "ELB지지대",
        "INSULATOR",
        "잡자재비",
        "ASSEMBLY CHARGE",
    ]

    def __init__(self):
        """초기화"""
        pass

    def validate(self, excel_path: Path, validator_report: ValidationReport) -> dict:
        """
        최종 문서 품질 검증

        Args:
            excel_path: 생성된 Excel 파일 경로
            validator_report: Stage 3 validator 결과

        Returns:
            {
                "passed": bool,  # 전체 통과 여부
                "score": float,  # 품질 점수 (0.0 ~ 1.0)
                "errors": List[str],  # 오류 목록
                "warnings": List[str],  # 경고 목록
                "checks": {
                    "formula_preservation": bool,
                    "cross_reference": bool,
                    "required_items": bool,
                },
            }
        """
        errors = []
        warnings = []
        checks = {
            "formula_preservation": False,
            "cross_reference": False,
            "required_items": False,
        }

        # 1. Stage 3 validator 결과 확인
        if not validator_report.formula_preservation:
            errors.append("수식 보존 실패 (CHK_FORMULA_PRESERVATION)")
        else:
            checks["formula_preservation"] = True

        if not validator_report.cross_references_valid:
            errors.append("크로스 참조 오류 (CHK_CROSS_SHEET_REFERENCE)")
        else:
            checks["cross_reference"] = True

        # 2. Excel 파일 열기 (실물 체크, 목업 금지)
        if not excel_path.exists():
            errors.append(f"Excel 파일 없음: {excel_path}")
            return {
                "passed": False,
                "score": 0.0,
                "errors": errors,
                "warnings": warnings,
                "checks": checks,
            }

        try:
            wb = load_workbook(excel_path, data_only=False)
            ws_estimate = wb["견적서"]
        except Exception as e:
            errors.append(f"Excel 파일 열기 실패: {e}")
            return {
                "passed": False,
                "score": 0.0,
                "errors": errors,
                "warnings": warnings,
                "checks": checks,
            }

        # 3. 필수 자재 12개 누락 체크
        required_items_check = self._check_required_items(ws_estimate)
        if required_items_check["missing_items"]:
            errors.append(
                f"필수 자재 누락 (견적조건 7번): {', '.join(required_items_check['missing_items'])}"
            )
        else:
            checks["required_items"] = True

        if required_items_check["warnings"]:
            warnings.extend(required_items_check["warnings"])

        # 4. 품질 점수 계산
        total_checks = len(checks)
        passed_checks = sum(1 for v in checks.values() if v)
        score = passed_checks / total_checks if total_checks > 0 else 0.0

        # 5. 전체 통과 여부
        passed = len(errors) == 0

        logger.info(
            f"Doc Lint: passed={passed}, score={score:.2f}, "
            f"errors={len(errors)}, warnings={len(warnings)}"
        )

        return {
            "passed": passed,
            "score": score,
            "errors": errors,
            "warnings": warnings,
            "checks": checks,
        }

    def _check_required_items(self, ws: Worksheet) -> dict:
        """
        견적조건 7번: 필수 자재 12개 누락 체크

        Args:
            ws: 견적서 시트

        Returns:
            {
                "missing_items": List[str],  # 누락된 항목
                "warnings": List[str],  # 경고
            }
        """
        missing_items = []
        warnings = []

        # B 컬럼 (품명)에서 필수 자재 찾기
        found_items = set()
        for row in ws.iter_rows(min_row=2, max_row=100, min_col=2, max_col=2):
            cell_value = row[0].value
            if cell_value:
                item_name = str(cell_value).strip()
                # 필수 자재 체크
                for required_item in self.REQUIRED_ITEMS:
                    if required_item in item_name:
                        found_items.add(required_item)

        # 누락된 항목 확인
        for required_item in self.REQUIRED_ITEMS:
            if required_item not in found_items:
                # 예외 처리: 일부 항목은 조건부
                if required_item == "차단기지지대":
                    # 400AF 이상만 필요 (경고만)
                    warnings.append("차단기지지대 없음 (400AF 미만이면 정상)")
                elif required_item == "BUS-BAR":
                    # 분기 있거나 부스바 처리 옵션일 때만
                    warnings.append("BUS-BAR 없음 (메인만 + 400AF 미만이면 정상)")
                elif required_item == "ELB지지대":
                    # 소형차단기(SIE-32, SIB-32, 32GRHS, BS-32)가 있을 때만 필요
                    warnings.append("ELB지지대 없음 (소형차단기 미포함이면 정상)")
                else:
                    missing_items.append(required_item)

        return {
            "missing_items": missing_items,
            "warnings": warnings,
        }
