"""
Excel 서식 스타일 정의 모듈

실물 템플릿(견적서양식.xlsx) 기반 스타일 추출
- Font: 맑은고딕 12pt
- Alignment: center/left/right
- Border: thin 실선
- Fill: 회색(D9D9D9)

사용법:
    from excel_styles import apply_data_row_style, apply_subtotal_style
    apply_data_row_style(ws, row_idx)
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ============================================================
# 기본 스타일 정의 (실물 템플릿 기준)
# ============================================================

# Font
DEFAULT_FONT = Font(name="맑은고딕", size=12, bold=False)
BOLD_FONT = Font(name="맑은고딕", size=12, bold=True)

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Border
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Fill
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)


# ============================================================
# 행 스타일 적용 함수
# ============================================================


def apply_data_row_style(ws, row_idx: int):
    """
    데이터 행 스타일 적용

    실물 템플릿 Row 3~47 패턴:
    - A: 순번 (center)
    - B: 품명 (left)
    - C: 규격 (left)
    - D: 단위 (center)
    - E: 수량 (right)
    - F: 단가 (right)
    - G: 금액 (수식, right)

    Args:
        ws: Worksheet
        row_idx: 행 번호
    """
    # A열: 순번 (center)
    ws[f"A{row_idx}"].font = DEFAULT_FONT
    ws[f"A{row_idx}"].alignment = ALIGN_CENTER
    ws[f"A{row_idx}"].border = THIN_BORDER

    # B열: 품명 (left)
    ws[f"B{row_idx}"].font = DEFAULT_FONT
    ws[f"B{row_idx}"].alignment = ALIGN_LEFT
    ws[f"B{row_idx}"].border = THIN_BORDER

    # C열: 규격 (left)
    ws[f"C{row_idx}"].font = DEFAULT_FONT
    ws[f"C{row_idx}"].alignment = ALIGN_LEFT
    ws[f"C{row_idx}"].border = THIN_BORDER

    # D열: 단위 (center)
    ws[f"D{row_idx}"].font = DEFAULT_FONT
    ws[f"D{row_idx}"].alignment = ALIGN_CENTER
    ws[f"D{row_idx}"].border = THIN_BORDER

    # E열: 수량 (right)
    ws[f"E{row_idx}"].font = DEFAULT_FONT
    ws[f"E{row_idx}"].alignment = ALIGN_RIGHT
    ws[f"E{row_idx}"].border = THIN_BORDER

    # F열: 단가 (right)
    ws[f"F{row_idx}"].font = DEFAULT_FONT
    ws[f"F{row_idx}"].alignment = ALIGN_RIGHT
    ws[f"F{row_idx}"].border = THIN_BORDER

    # G열: 금액 (right, 수식 보존)
    ws[f"G{row_idx}"].font = DEFAULT_FONT
    ws[f"G{row_idx}"].alignment = ALIGN_RIGHT
    ws[f"G{row_idx}"].border = THIN_BORDER


def apply_subtotal_style(ws, row_idx: int):
    """
    소계/합계 행 스타일 적용

    실물 템플릿 Row 48~49 패턴:
    - A~F: 회색 배경 (D9D9D9)
    - G: 금액 (수식, 회색 배경)

    Args:
        ws: Worksheet
        row_idx: 행 번호
    """
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        cell = ws[f"{col}{row_idx}"]
        cell.font = BOLD_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        cell.fill = GRAY_FILL


def apply_header_style(ws, row_idx: int):
    """
    헤더 행 스타일 적용

    실물 템플릿 Row 2 패턴:
    - 모든 셀: bold + center + 회색 배경

    Args:
        ws: Worksheet
        row_idx: 행 번호
    """
    header_labels = ["순번", "품명", "규격", "단위", "수량", "단가", "금액"]

    for idx, label in enumerate(header_labels, start=1):
        col_letter = chr(64 + idx)  # A=65, B=66, ...
        cell = ws[f"{col_letter}{row_idx}"]
        cell.value = label
        cell.font = BOLD_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        cell.fill = GRAY_FILL


def apply_panel_name_style(ws, row_idx: int, panel_name: str):
    """
    분전반 이름 행 스타일 적용

    패턴:
    - A열: 분전반 이름 (bold + left)
    - B~G: 병합

    Args:
        ws: Worksheet
        row_idx: 행 번호
        panel_name: 분전반 이름
    """
    ws[f"A{row_idx}"].value = f"■ {panel_name}"
    ws[f"A{row_idx}"].font = BOLD_FONT
    ws[f"A{row_idx}"].alignment = ALIGN_LEFT
    ws[f"A{row_idx}"].border = THIN_BORDER

    # 병합 셀 (A~G)
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
